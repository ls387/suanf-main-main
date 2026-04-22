# -*- coding: utf-8 -*-
"""
教师黑名单时间对话式收集器。

职责：
  1. 判断用户消息是否触发"黑名单收集"意图
  2. 通过多轮对话向用户追问缺失字段（学期、教师、星期、节次、原因）
  3. 信息收集完毕后生成 Excel 文件（bytes），并通过 SSE 事件通知前端下载
"""
import io
import re
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 星期映射
WEEKDAY_MAP = {
    "周一": 1, "周二": 2, "周三": 3, "周四": 4, "周五": 5, "周六": 6, "周日": 7,
    "星期一": 1, "星期二": 2, "星期三": 3, "星期四": 4, "星期五": 5, "星期六": 6, "星期日": 7,
    "一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "日": 7,
}

SLOT_PATTERN = re.compile(
    r"第?\s*(\d{1,2})\s*[节\-–—~～到至]\s*第?\s*(\d{1,2})\s*节?|第?\s*(\d{1,2})\s*节"
)


def _extract_weekday(text: str) -> Optional[int]:
    for k, v in WEEKDAY_MAP.items():
        if k in text:
            return v
    m = re.search(r"周(\d)", text)
    if m:
        return int(m.group(1))
    return None


def _extract_slots(text: str) -> tuple[Optional[int], Optional[int]]:
    m = SLOT_PATTERN.search(text)
    if not m:
        return None, None
    if m.group(1) and m.group(2):
        return int(m.group(1)), int(m.group(2))
    if m.group(3):
        s = int(m.group(3))
        return s, s
    return None, None


def _extract_semester(text: str) -> Optional[str]:
    m = re.search(r"(\d{4})-(\d{4})-([12])", text)
    if m:
        return m.group(0)
    m = re.search(r"(\d{2})-(\d{2})-([12])", text)
    if m:
        g = m.group(0).split("-")
        return f"20{g[0]}-20{g[1]}-{g[2]}"
    now = datetime.now()
    year, month = now.year, now.month
    if "下学期" in text or "下半学期" in text:
        if month >= 8:
            return f"{year}-{year+1}-2"
        return f"{year-1}-{year}-2"
    if "上学期" in text or "上半学期" in text:
        if month >= 8:
            return f"{year}-{year+1}-1"
        return f"{year-1}-{year}-1"
    return None


def _extract_teacher_id(text: str) -> Optional[str]:
    m = re.search(r"\b([A-Z]{2,8}\d{3,6})\b", text)
    if m:
        return m.group(1)
    return None


# ── 意图检测 ──────────────────────────────────────────────────────────────────

BLACKOUT_TRIGGER_KEYWORDS = [
    "不想上课", "不能上课", "不方便上课", "不想在", "不能在",
    "黑名单", "不可用时间", "屏蔽时间", "禁止排课", "避开",
    "不要排在", "不安排", "空出", "留出",
]


def is_blackout_intent(text: str) -> bool:
    return any(kw in text for kw in BLACKOUT_TRIGGER_KEYWORDS)


# ── 状态机 ────────────────────────────────────────────────────────────────────

class BlackoutSession:
    __slots__ = ("teacher_id", "teacher_name", "semester", "weekday",
                 "start_slot", "end_slot", "reason", "confirmed")

    def __init__(self):
        self.teacher_id: Optional[str] = None
        self.teacher_name: Optional[str] = None
        self.semester: Optional[str] = None
        self.weekday: Optional[int] = None
        self.start_slot: Optional[int] = None
        self.end_slot: Optional[int] = None
        self.reason: Optional[str] = None
        self.confirmed: bool = False

    def is_complete(self) -> bool:
        return all([
            self.teacher_id,
            self.semester,
            self.weekday is not None,
            self.start_slot is not None,
            self.end_slot is not None,
        ])

    def to_dict(self) -> dict:
        return {
            "teacher_id": self.teacher_id,
            "semester": self.semester,
            "weekday": self.weekday,
            "start_slot": self.start_slot,
            "end_slot": self.end_slot,
            "reason": self.reason or "",
        }


_blackout_sessions: dict[str, Optional[BlackoutSession]] = {}


def get_blackout_session(session_id: str) -> Optional[BlackoutSession]:
    return _blackout_sessions.get(session_id)


def start_blackout_session(session_id: str) -> BlackoutSession:
    bs = BlackoutSession()
    _blackout_sessions[session_id] = bs
    return bs


def end_blackout_session(session_id: str):
    _blackout_sessions.pop(session_id, None)


# ── 主处理函数 ────────────────────────────────────────────────────────────────

def process_blackout_turn(session_id: str, user_text: str) -> dict:
    """
    处理一轮用户输入，返回 dict：
      {"reply": str, "done": bool, "excel_key": str|None}
    excel_key 非空时表示 Excel 已生成，前端可凭此 key 下载。
    """
    bs = _blackout_sessions.get(session_id)
    if bs is None:
        bs = start_blackout_session(session_id)

    _parse_into(bs, user_text)

    if any(w in user_text for w in ["取消", "算了", "不用了", "退出"]):
        end_blackout_session(session_id)
        return {
            "reply": "好的，已取消黑名单时间录入。有其他问题随时告诉我！",
            "done": True,
            "excel_key": None,
        }

    if bs.is_complete() and any(w in user_text for w in ["确认", "是", "对", "没错", "OK", "ok", "好的", "确定"]):
        bs.confirmed = True

    if bs.confirmed and bs.is_complete():
        excel_bytes = _build_excel([bs.to_dict()])
        key = _store_excel(session_id, excel_bytes)
        name_part = f"（{bs.teacher_name}）" if bs.teacher_name else ""
        reason_part = bs.reason or "无"
        reply = (
            f"好的！已为您生成黑名单时间导入表：\n"
            f"- 教师ID：{bs.teacher_id}{name_part}\n"
            f"- 学期：{bs.semester}\n"
            f"- {_weekday_str(bs.weekday)} 第{bs.start_slot}—{bs.end_slot}节\n"
            f"- 原因：{reason_part}\n\n"
            "请点击下方按钮下载 Excel，然后前往【教师偏好与黑名单时间管理】页面录入。"
        )
        end_blackout_session(session_id)
        return {"reply": reply, "done": True, "excel_key": key}

    if bs.is_complete() and not bs.confirmed:
        name_part = f"（{bs.teacher_name}）" if bs.teacher_name else ""
        reason_part = bs.reason or "无"
        summary = (
            f"我帮您整理了以下黑名单时间，请确认：\n"
            f"- 教师ID：{bs.teacher_id}{name_part}\n"
            f"- 学期：{bs.semester}\n"
            f"- {_weekday_str(bs.weekday)} 第{bs.start_slot}—{bs.end_slot}节\n"
            f"- 原因：{reason_part}\n\n"
            '信息是否正确？请回复"确认"生成 Excel，或告诉我需要修改的地方。'
        )
        return {"reply": summary, "done": False, "excel_key": None}

    question = _next_question(bs)
    return {"reply": question, "done": False, "excel_key": None}


def _parse_into(bs: BlackoutSession, text: str):
    if not bs.teacher_id:
        tid = _extract_teacher_id(text)
        if tid:
            bs.teacher_id = tid
    if not bs.teacher_name:
        m = re.search(r"([一-龥]{2,4})老师", text)
        if m:
            bs.teacher_name = m.group(1)
    if not bs.semester:
        bs.semester = _extract_semester(text)
    if bs.weekday is None:
        bs.weekday = _extract_weekday(text)
    if bs.start_slot is None or bs.end_slot is None:
        s, e = _extract_slots(text)
        if s is not None:
            bs.start_slot = s
        if e is not None:
            bs.end_slot = e
    if not bs.reason:
        for kw in ["因为", "原因是", "由于"]:
            if kw in text:
                idx = text.index(kw) + len(kw)
                bs.reason = text[idx:].strip().rstrip("。，！")[:50]
                break


def _next_question(bs: BlackoutSession) -> str:
    if not bs.teacher_id:
        hint = f"（{bs.teacher_name}老师）" if bs.teacher_name else ""
        return f"请提供教师的工号/编号{hint}，例如 ISCIT106。"
    if not bs.semester:
        return "请告诉我是哪个学期？格式如 2025-2026-1（第一学期）或 2025-2026-2（第二学期）。"
    if bs.weekday is None:
        return "请问是星期几不方便上课？"
    if bs.start_slot is None or bs.end_slot is None:
        return '请问是第几节到第几节？例如"第1-2节"或"第3节"。'
    return "信息已基本收集完毕，请确认。"


def _weekday_str(wd: int) -> str:
    return ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][wd - 1]


# ── Excel 生成 ────────────────────────────────────────────────────────────────

def _build_excel(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "教师黑名单时间"

    headers = ["教师ID", "学期", "星期(数字)", "开始节次", "结束节次", "原因"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, 2):
        ws.cell(r, 1, row["teacher_id"])
        ws.cell(r, 2, row["semester"])
        ws.cell(r, 3, row["weekday"])
        ws.cell(r, 4, row["start_slot"])
        ws.cell(r, 5, row["end_slot"])
        ws.cell(r, 6, row["reason"])

    note_row = len(rows) + 3
    note_cell = ws.cell(note_row, 1, "说明：星期数字 1=周一，7=周日；节次范围 1–13")
    note_cell.font = Font(italic=True, color="808080")

    for col, width in zip("ABCDEF", [18, 16, 12, 12, 12, 30]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── 临时存储 ──────────────────────────────────────────────────────────────────

_excel_store: dict[str, bytes] = {}


def _store_excel(session_id: str, data: bytes) -> str:
    key = f"blackout_{session_id}_{int(datetime.now().timestamp())}"
    _excel_store[key] = data
    return key


def pop_excel(key: str) -> Optional[bytes]:
    return _excel_store.pop(key, None)
