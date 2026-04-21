# -*- coding: utf-8 -*-
"""
冲突检测路由

GET  /api/conflicts/{version_id}          完整冲突详情
GET  /api/conflicts/{version_id}/summary  仅 summary（轻量）
POST /api/conflicts/{version_id}/fix      修复指定类型冲突
GET  /api/conflicts/{version_id}/export   导出冲突 Excel
"""
import io
from datetime import datetime
from typing import Literal

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.services.conflict_analyzer import analyze_conflicts
from app.services.conflict_optimizer import fix_conflicts

router = APIRouter()

_DAY_NAMES = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_CONFLICT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


class FixRequest(BaseModel):
    fix_type: Literal["capacity", "class", "teacher", "classroom"]


@router.get("/{version_id}/summary")
async def get_conflict_summary(version_id: int):
    result = analyze_conflicts(version_id)
    return {"version_id": version_id, "summary": result["summary"]}


@router.get("/{version_id}/export")
async def export_conflicts(version_id: int):
    """导出冲突报告为 Excel 文件（流式下载）。"""
    data = analyze_conflicts(version_id)

    wb = Workbook()

    def _header(ws, cols):
        ws.append(cols)
        for i in range(1, len(cols) + 1):
            cell = ws.cell(1, i)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _time_range(start, end):
        return f"{start}-{end}节" if start != end else f"{start}节"

    # 班级冲突
    ws1 = wb.active
    ws1.title = "班级冲突"
    _header(ws1, ["序号", "班级", "星期", "冲突节次", "课程1", "教师1", "教室1", "课程2", "教师2", "教室2"])
    for idx, c in enumerate(data["class_conflicts"], 1):
        ws1.append([
            idx,
            c.get("class_name", ""),
            _DAY_NAMES[c["weekday"]],
            _time_range(c["overlap_start"], c["overlap_end"]),
            c.get("course1", ""), c.get("teacher1", ""), c.get("classroom1", ""),
            c.get("course2", ""), c.get("teacher2", ""), c.get("classroom2", ""),
        ])
        for col in range(1, 11):
            ws1.cell(idx + 1, col).fill = _CONFLICT_FILL
    for col_letter, width in zip("ABCDEFGHIJ", [8, 20, 8, 10, 20, 15, 15, 20, 15, 15]):
        ws1.column_dimensions[col_letter].width = width

    # 教师冲突
    ws2 = wb.create_sheet("教师冲突")
    _header(ws2, ["序号", "教师", "星期", "冲突节次", "课程1", "教室1", "课程2", "教室2"])
    for idx, c in enumerate(data["teacher_conflicts"], 1):
        ws2.append([
            idx,
            c.get("teacher", ""),
            _DAY_NAMES[c["weekday"]],
            _time_range(c["overlap_start"], c["overlap_end"]),
            c.get("course1", ""), c.get("classroom1", ""),
            c.get("course2", ""), c.get("classroom2", ""),
        ])
        for col in range(1, 9):
            ws2.cell(idx + 1, col).fill = _CONFLICT_FILL
    for col_letter, width in zip("ABCDEFGH", [8, 15, 8, 10, 20, 15, 20, 15]):
        ws2.column_dimensions[col_letter].width = width

    # 教室冲突
    ws3 = wb.create_sheet("教室冲突")
    _header(ws3, ["序号", "教室", "星期", "冲突节次", "课程1", "教师1", "课程2", "教师2"])
    for idx, c in enumerate(data["classroom_conflicts"], 1):
        ws3.append([
            idx,
            c.get("classroom", ""),
            _DAY_NAMES[c["weekday"]],
            _time_range(c["overlap_start"], c["overlap_end"]),
            c.get("course1", ""), c.get("teacher1", ""),
            c.get("course2", ""), c.get("teacher2", ""),
        ])
        for col in range(1, 9):
            ws3.cell(idx + 1, col).fill = _CONFLICT_FILL
    for col_letter, width in zip("ABCDEFGH", [8, 20, 8, 10, 20, 15, 20, 15]):
        ws3.column_dimensions[col_letter].width = width

    # 容量不足
    ws4 = wb.create_sheet("容量不足")
    _header(ws4, ["序号", "课程", "星期", "节次", "教室", "教室容量", "学生数", "缺座"])
    for idx, c in enumerate(data["capacity_violations"], 1):
        ws4.append([
            idx,
            c.get("course", ""),
            _DAY_NAMES[c["weekday"]],
            _time_range(c["start_slot"], c["end_slot"]),
            c.get("classroom", ""),
            c.get("capacity", ""),
            c.get("students", ""),
            c.get("shortage", ""),
        ])
        for col in range(1, 9):
            ws4.cell(idx + 1, col).fill = _CONFLICT_FILL
    for col_letter, width in zip("ABCDEFGH", [8, 20, 8, 10, 15, 10, 10, 8]):
        ws4.column_dimensions[col_letter].width = width

    # 输出为字节流
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"排课冲突报告_版本{version_id}_{timestamp}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{version_id}/fix")
async def fix_version_conflicts(version_id: int, body: FixRequest):
    """修复指定类型的冲突，直接更新数据库。"""
    try:
        result = fix_conflicts(version_id, body.fix_type)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return result


@router.get("/{version_id}")
async def get_conflicts(version_id: int):
    return analyze_conflicts(version_id)
