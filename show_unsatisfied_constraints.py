#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
显示未满足的约束和优化建议
分析排课结果中未达到最优的软约束，并提供优化建议
"""
import os
import sys
import pymysql
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 设置标准输出编码为UTF-8
if sys.platform == "win32":
    import io

    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


def get_db_connection():
    """获取数据库连接"""
    return pymysql.connect(
        host=os.getenv("DB_HOST") or "localhost",
        port=int(os.getenv("DB_PORT") or "3306"),
        user=os.getenv("DB_USER") or "pk",
        password=os.getenv("DB_PASSWORD") or "123456",
        database=os.getenv("DB_NAME") or "paike",
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
    )


def analyze_unsatisfied_constraints(version_id):
    """分析未满足的约束"""
    conn = get_db_connection()
    cursor = conn.cursor()

    print(f"\n{'='*70}")
    print(f"排课方案未满足约束分析 - 版本 {version_id}")
    print(f"{'='*70}\n")

    # 1. 教师偏好未满足
    preference_violations = analyze_teacher_preferences(version_id, cursor)

    # 2. 教室利用率问题
    utilization_issues = analyze_classroom_utilization(version_id, cursor)

    # 3. 学生负荷问题
    student_overload = analyze_student_overload(version_id, cursor)

    # 4. 任务关系约束未满足
    task_relation_violations = analyze_task_relations(version_id, cursor)

    # 5. 课程时段偏好问题
    time_preference_issues = analyze_time_preferences(version_id, cursor)

    # 6. 教室连续性问题
    continuity_issues = analyze_classroom_continuity(version_id, cursor)

    # 7. 校区通勤问题
    commute_issues = analyze_campus_commute(version_id, cursor)

    # 生成综合报告
    generate_comprehensive_report(
        version_id,
        preference_violations,
        utilization_issues,
        student_overload,
        task_relation_violations,
        time_preference_issues,
        continuity_issues,
        commute_issues,
    )

    cursor.close()
    conn.close()


def analyze_teacher_preferences(version_id, cursor):
    """分析教师偏好未满足情况"""
    print("\n【1. 教师个性化偏好分析】")
    print("-" * 70)

    # 获取排课结果
    query = """
    SELECT 
        s.schedule_id,
        s.task_id,
        s.week_day,
        s.start_slot,
        tt.slots_count,
        c.course_name,
        t.teacher_id,
        t.teacher_name,
        cr.classroom_name
    FROM schedules s
    JOIN teaching_tasks tt ON s.task_id = tt.task_id
    JOIN course_offerings co ON tt.offering_id = co.offering_id
    JOIN courses c ON co.course_id = c.course_id
    JOIN offering_teachers ot ON co.offering_id = ot.offering_id
    JOIN teachers t ON ot.teacher_id = t.teacher_id
    JOIN classrooms cr ON s.classroom_id = cr.classroom_id
    WHERE s.version_id = %s
    """
    cursor.execute(query, (version_id,))
    schedules = cursor.fetchall()

    # 构建教师排课映射
    teacher_schedule = defaultdict(list)
    for sch in schedules:
        teacher_schedule[sch["teacher_id"]].append(sch)

    # 获取教师偏好
    pref_query = """
    SELECT 
        tp.*,
        t.teacher_name,
        c.course_name
    FROM teacher_preferences tp
    JOIN teachers t ON tp.teacher_id = t.teacher_id
    JOIN course_offerings co ON tp.offering_id = co.offering_id
    JOIN courses c ON co.course_id = c.course_id
    WHERE tp.offering_id IN (
        SELECT DISTINCT tt.offering_id 
        FROM teaching_tasks tt
        JOIN schedules s ON tt.task_id = s.task_id
        WHERE s.version_id = %s
    )
    ORDER BY tp.penalty_score DESC
    """
    cursor.execute(pref_query, (version_id,))
    preferences = cursor.fetchall()

    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    violations = {"avoided": [], "preferred": []}  # 违反了避免时段  # 未在偏好时段

    for pref in preferences:
        teacher_id = pref["teacher_id"]
        pref_type = pref["preference_type"]
        weekday = pref["weekday"]
        start_slot = pref["start_slot"]
        end_slot = pref["end_slot"]

        if not weekday or not start_slot or not end_slot:
            continue

        teacher_courses = teacher_schedule.get(teacher_id, [])

        for course in teacher_courses:
            course_weekday = course["week_day"]
            course_start = course["start_slot"]
            course_end = course["start_slot"] + course["slots_count"] - 1

            # 检查时间重叠
            if course_weekday == weekday:
                overlap = not (course_end < start_slot or course_start > end_slot)

                if pref_type == "AVOIDED" and overlap:
                    violations["avoided"].append(
                        {
                            "teacher_name": pref["teacher_name"],
                            "course_name": course["course_name"],
                            "weekday": day_names[course_weekday],
                            "time": f"{course_start}-{course_end}节",
                            "avoid_time": f"{day_names[weekday]} {start_slot}-{end_slot}节",
                            "penalty": pref["penalty_score"],
                            "classroom": course["classroom_name"],
                        }
                    )
                elif pref_type == "PREFERRED" and not overlap:
                    # 有偏好但不在偏好时段
                    pass  # 这个需要更复杂的判断

    # 显示结果
    if violations["avoided"]:
        print(f"\n⚠️  违反避免时段: {len(violations['avoided'])} 处")
        for i, v in enumerate(violations["avoided"][:10], 1):  # 只显示前10个
            print(f"  {i}. {v['teacher_name']} - {v['course_name']}")
            print(f"     实际: {v['weekday']} {v['time']} ({v['classroom']})")
            print(f"     避免: {v['avoid_time']}")
            print(f"     惩罚分数: {v['penalty']}")

        if len(violations["avoided"]) > 10:
            print(f"  ... 还有 {len(violations['avoided']) - 10} 处未显示")
    else:
        print("✅ 所有教师避免时段均已满足")

    return violations


def analyze_classroom_utilization(version_id, cursor):
    """分析教室利用率问题"""
    print("\n【2. 教室容量利用率分析】")
    print("-" * 70)

    query = """
    SELECT 
        s.schedule_id,
        s.week_day,
        s.start_slot,
        tt.slots_count,
        c.course_name,
        cr.classroom_name,
        cr.capacity,
        co.student_count_estimate
    FROM schedules s
    JOIN teaching_tasks tt ON s.task_id = tt.task_id
    JOIN course_offerings co ON tt.offering_id = co.offering_id
    JOIN courses c ON co.course_id = c.course_id
    JOIN classrooms cr ON s.classroom_id = cr.classroom_id
    WHERE s.version_id = %s AND cr.capacity > 0
    """
    cursor.execute(query, (version_id,))
    results = cursor.fetchall()

    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    # 分析利用率
    waste_issues = []  # 利用率<50%
    tight_issues = []  # 利用率90%-100%

    for row in results:
        capacity = row["capacity"]
        students = row["student_count_estimate"] or 0

        if students == 0:
            continue

        utilization = students / capacity if capacity > 0 else 0

        if utilization < 0.5:
            waste_issues.append(
                {
                    "course": row["course_name"],
                    "classroom": row["classroom_name"],
                    "capacity": capacity,
                    "students": students,
                    "utilization": utilization * 100,
                    "weekday": day_names[row["week_day"]],
                    "time": f"{row['start_slot']}-{row['start_slot'] + row['slots_count'] - 1}节",
                    "waste_seats": capacity - students,
                }
            )
        elif utilization >= 0.9:
            tight_issues.append(
                {
                    "course": row["course_name"],
                    "classroom": row["classroom_name"],
                    "capacity": capacity,
                    "students": students,
                    "utilization": utilization * 100,
                    "weekday": day_names[row["week_day"]],
                    "time": f"{row['start_slot']}-{row['start_slot'] + row['slots_count'] - 1}节",
                }
            )

    # 按浪费座位数排序
    waste_issues.sort(key=lambda x: x["waste_seats"], reverse=True)

    if waste_issues:
        print(f"\n⚠️  教室利用率低(<50%): {len(waste_issues)} 处")
        print(f"  (建议更换为容量更小的教室)")
        for i, issue in enumerate(waste_issues[:10], 1):
            print(f"  {i}. {issue['course']} - {issue['classroom']}")
            print(f"     {issue['weekday']} {issue['time']}")
            print(
                f"     容量:{issue['capacity']} 学生:{issue['students']} "
                f"利用率:{issue['utilization']:.1f}% 浪费:{issue['waste_seats']}座"
            )

        if len(waste_issues) > 10:
            print(f"  ... 还有 {len(waste_issues) - 10} 处未显示")
    else:
        print("✅ 所有教室利用率均在合理范围")

    if tight_issues:
        print(f"\n💡 教室较为拥挤(≥90%): {len(tight_issues)} 处")
        for i, issue in enumerate(tight_issues[:5], 1):
            print(f"  {i}. {issue['course']} - {issue['classroom']}")
            print(
                f"     容量:{issue['capacity']} 学生:{issue['students']} "
                f"利用率:{issue['utilization']:.1f}%"
            )

    return {"waste": waste_issues, "tight": tight_issues}


def analyze_student_overload(version_id, cursor):
    """分析学生每日课时负荷"""
    print("\n【3. 学生每日课时负荷分析】")
    print("-" * 70)

    query = """
    SELECT 
        s.week_day,
        s.start_slot,
        tt.slots_count,
        c.course_name,
        cl.class_id,
        cl.class_name
    FROM schedules s
    JOIN teaching_tasks tt ON s.task_id = tt.task_id
    JOIN course_offerings co ON tt.offering_id = co.offering_id
    JOIN courses c ON co.course_id = c.course_id
    JOIN offering_classes oc ON co.offering_id = oc.offering_id
    JOIN classes cl ON oc.class_id = cl.class_id
    WHERE s.version_id = %s
    ORDER BY cl.class_id, s.week_day, s.start_slot
    """
    cursor.execute(query, (version_id,))
    results = cursor.fetchall()

    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    # 统计每个班级每天的课时数
    class_daily_load = defaultdict(lambda: defaultdict(list))

    for row in results:
        class_id = row["class_id"]
        weekday = row["week_day"]
        class_daily_load[class_id][weekday].append(
            {
                "course": row["course_name"],
                "slots": row["slots_count"],
                "start": row["start_slot"],
            }
        )

    overload_issues = []

    for class_id, daily_load in class_daily_load.items():
        for weekday, courses in daily_load.items():
            total_slots = sum(c["slots"] for c in courses)

            if total_slots >= 8:  # 一天8节或以上算过载
                # 获取班级名称
                class_name = next(
                    (c["class_name"] for c in results if c["class_id"] == class_id),
                    class_id,
                )

                overload_issues.append(
                    {
                        "class_name": class_name,
                        "weekday": day_names[weekday],
                        "total_slots": total_slots,
                        "courses": [c["course"] for c in courses],
                        "course_count": len(courses),
                    }
                )

    overload_issues.sort(key=lambda x: x["total_slots"], reverse=True)

    if overload_issues:
        print(f"\n⚠️  学生每日课时过多(≥8节): {len(overload_issues)} 处")
        for i, issue in enumerate(overload_issues[:10], 1):
            print(f"  {i}. {issue['class_name']} - {issue['weekday']}")
            print(f"     共{issue['total_slots']}节，{issue['course_count']}门课程")
            print(f"     课程: {', '.join(issue['courses'][:3])}")

        if len(overload_issues) > 10:
            print(f"  ... 还有 {len(overload_issues) - 10} 处未显示")
    else:
        print("✅ 所有班级每日课时负荷均在合理范围")

    return overload_issues


def analyze_task_relations(version_id, cursor):
    """分析任务关系约束未满足情况"""
    print("\n【4. 课程任务关系约束分析】")
    print("-" * 70)

    try:
        # 获取任务关系约束
        query = """
        SELECT 
            trc.*,
            tta.task_id AS task_id_a,
            ttb.task_id AS task_id_b,
            c.course_name
        FROM task_relation_constraints trc
        JOIN course_offerings co ON trc.offering_id = co.offering_id
        JOIN courses c ON co.course_id = c.course_id
        JOIN teaching_tasks tta ON co.offering_id = tta.offering_id 
            AND tta.task_sequence = trc.task_sequence_a
        JOIN teaching_tasks ttb ON co.offering_id = ttb.offering_id 
            AND ttb.task_sequence = trc.task_sequence_b
        WHERE co.offering_id IN (
            SELECT DISTINCT tt.offering_id 
            FROM teaching_tasks tt
            JOIN schedules s ON tt.task_id = s.task_id
            WHERE s.version_id = %s
        )
        """
        cursor.execute(query, (version_id,))
        constraints = cursor.fetchall()

        if not constraints:
            print("✅ 未设置任务关系约束")
            return []

        violations = []
        day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

        for const in constraints:
            # 获取两个任务的排课信息
            cursor.execute(
                "SELECT task_id, week_day, start_slot FROM schedules WHERE version_id = %s AND task_id = %s",
                (version_id, const["task_id_a"]),
            )
            sch_a = cursor.fetchone()

            cursor.execute(
                "SELECT task_id, week_day, start_slot FROM schedules WHERE version_id = %s AND task_id = %s",
                (version_id, const["task_id_b"]),
            )
            sch_b = cursor.fetchone()

            if not sch_a or not sch_b:
                continue

            day_diff = abs(sch_a["week_day"] - sch_b["week_day"])

            constraint_type = const["constraint_type"]
            violated = False
            reason = ""

            if constraint_type == "REQUIRE_SAME_DAY":
                if sch_a["week_day"] != sch_b["week_day"]:
                    violated = True
                    reason = f"要求同一天，实际: {day_names[sch_a['week_day']]} 和 {day_names[sch_b['week_day']]}"

            elif constraint_type == "AVOID_CONSECUTIVE_DAYS":
                if day_diff == 1:
                    violated = True
                    reason = f"要求避免连续天，实际: {day_names[sch_a['week_day']]} 和 {day_names[sch_b['week_day']]}"

            elif constraint_type == "MIN_DAYS_APART":
                min_gap = const["constraint_value"] or 1
                if day_diff < min_gap:
                    violated = True
                    reason = f"要求至少间隔{min_gap}天，实际间隔{day_diff}天"

            if violated:
                violations.append(
                    {
                        "course": const["course_name"],
                        "constraint_type": constraint_type,
                        "reason": reason,
                        "task_a_time": f"{day_names[sch_a['week_day']]} 第{sch_a['start_slot']}节",
                        "task_b_time": f"{day_names[sch_b['week_day']]} 第{sch_b['start_slot']}节",
                        "penalty": const["penalty_score"],
                        "task_id_a": const["task_id_a"],
                        "task_id_b": const["task_id_b"],
                        "can_optimize": True,  # 标记为可优化
                    }
                )

        if violations:
            print(f"\n⚠️  任务关系约束未满足: {len(violations)} 处")
            for i, v in enumerate(violations[:10], 1):
                print(f"  {i}. {v['course']} - {v['constraint_type']}")
                print(f"     {v['reason']}")
                print(f"     第1次: {v['task_a_time']}, 第2次: {v['task_b_time']}")
                print(f"     惩罚分数: {v['penalty']}")
                print(f"     💡 可通过调整时间优化")

            if len(violations) > 10:
                print(f"  ... 还有 {len(violations) - 10} 处未显示")
        else:
            print("✅ 所有任务关系约束均已满足")

        return violations

    except Exception as e:
        print(f"ℹ️  任务关系约束功能出错: {str(e)}")
        import traceback

        traceback.print_exc()
        return []


def analyze_time_preferences(version_id, cursor):
    """分析课程时段偏好（必修课避免晚上/周末）"""
    print("\n【5. 课程时段偏好分析】")
    print("-" * 70)

    query = """
    SELECT 
        s.week_day,
        s.start_slot,
        tt.slots_count,
        c.course_name,
        co.course_nature
    FROM schedules s
    JOIN teaching_tasks tt ON s.task_id = tt.task_id
    JOIN course_offerings co ON tt.offering_id = co.offering_id
    JOIN courses c ON co.course_id = c.course_id
    WHERE s.version_id = %s
    """
    cursor.execute(query, (version_id,))
    results = cursor.fetchall()

    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    night_issues = []  # 必修课在晚上
    weekend_issues = []  # 必修课在周末
    prime_time_issues = []  # 选修课占用黄金时段

    for row in results:
        weekday = row["week_day"]
        start_slot = row["start_slot"]
        end_slot = start_slot + row["slots_count"] - 1
        nature = row["course_nature"]

        # 检查必修/通识课在晚上(11-13节)
        if nature in ["必修", "通识"]:
            if start_slot >= 11:
                night_issues.append(
                    {
                        "course": row["course_name"],
                        "nature": nature,
                        "weekday": day_names[weekday],
                        "time": f"{start_slot}-{end_slot}节",
                    }
                )

            # 检查必修课在周末
            if weekday >= 6:
                weekend_issues.append(
                    {
                        "course": row["course_name"],
                        "nature": nature,
                        "weekday": day_names[weekday],
                        "time": f"{start_slot}-{end_slot}节",
                    }
                )

        # 检查选修课占用黄金时段(上午1-5，下午6-8)
        elif nature == "选修":
            if (start_slot >= 1 and start_slot <= 5) or (
                start_slot >= 6 and start_slot <= 8
            ):
                prime_time_issues.append(
                    {
                        "course": row["course_name"],
                        "weekday": day_names[weekday],
                        "time": f"{start_slot}-{end_slot}节",
                    }
                )

    if night_issues:
        print(f"\n⚠️  必修/通识课安排在晚上: {len(night_issues)} 处")
        for i, issue in enumerate(night_issues[:5], 1):
            print(f"  {i}. {issue['course']} ({issue['nature']})")
            print(f"     {issue['weekday']} {issue['time']}")

    if weekend_issues:
        print(f"\n⚠️  必修课安排在周末: {len(weekend_issues)} 处")
        for i, issue in enumerate(weekend_issues[:5], 1):
            print(f"  {i}. {issue['course']} ({issue['nature']})")
            print(f"     {issue['weekday']} {issue['time']}")

    if not night_issues and not weekend_issues:
        print("✅ 必修/通识课时段安排合理")

    return {
        "night": night_issues,
        "weekend": weekend_issues,
        "prime_time": prime_time_issues,
    }


def analyze_classroom_continuity(version_id, cursor):
    """分析同一课程是否在同一教室连续上课"""
    print("\n【6. 教室连续性分析】")
    print("-" * 70)

    query = """
    SELECT 
        tt.offering_id,
        c.course_name,
        COUNT(DISTINCT s.classroom_id) as classroom_count
    FROM schedules s
    JOIN teaching_tasks tt ON s.task_id = tt.task_id
    JOIN course_offerings co ON tt.offering_id = co.offering_id
    JOIN courses c ON co.course_id = c.course_id
    WHERE s.version_id = %s
    GROUP BY tt.offering_id, c.course_name
    HAVING classroom_count > 1
    """
    cursor.execute(query, (version_id,))
    results = cursor.fetchall()

    issues = []

    for row in results:
        # 获取该课程使用的所有教室
        detail_query = """
        SELECT cr.classroom_name, COUNT(*) as usage_count
        FROM schedules s
        JOIN teaching_tasks tt ON s.task_id = tt.task_id
        JOIN classrooms cr ON s.classroom_id = cr.classroom_id
        WHERE s.version_id = %s AND tt.offering_id = %s
        GROUP BY cr.classroom_name
        ORDER BY usage_count DESC
        """
        cursor.execute(detail_query, (version_id, row["offering_id"]))
        classrooms = cursor.fetchall()

        issues.append(
            {
                "course": row["course_name"],
                "classroom_count": row["classroom_count"],
                "classrooms": [
                    f"{c['classroom_name']}({c['usage_count']}次)" for c in classrooms
                ],
            }
        )

    if issues:
        print(f"\n💡 同一课程使用多个教室: {len(issues)} 门课程")
        for i, issue in enumerate(issues[:10], 1):
            print(f"  {i}. {issue['course']}")
            print(
                f"     使用了{issue['classroom_count']}个教室: {', '.join(issue['classrooms'][:3])}"
            )

        if len(issues) > 10:
            print(f"  ... 还有 {len(issues) - 10} 门课程未显示")
    else:
        print("✅ 所有课程均在固定教室上课")

    return issues


def analyze_campus_commute(version_id, cursor):
    """分析教师跨校区通勤问题"""
    print("\n【7. 教师跨校区通勤分析】")
    print("-" * 70)

    query = """
    SELECT 
        t.teacher_id,
        t.teacher_name,
        s.week_day,
        s.start_slot,
        cr.campus_id,
        ca.campus_name,
        c.course_name
    FROM schedules s
    JOIN teaching_tasks tt ON s.task_id = tt.task_id
    JOIN course_offerings co ON tt.offering_id = co.offering_id
    JOIN courses c ON co.course_id = c.course_id
    JOIN offering_teachers ot ON co.offering_id = ot.offering_id
    JOIN teachers t ON ot.teacher_id = t.teacher_id
    JOIN classrooms cr ON s.classroom_id = cr.classroom_id
    JOIN campuses ca ON cr.campus_id = ca.campus_id
    WHERE s.version_id = %s
    ORDER BY t.teacher_id, s.week_day, s.start_slot
    """
    cursor.execute(query, (version_id,))
    results = cursor.fetchall()

    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    # 按教师和日期分组
    teacher_daily_campus = defaultdict(lambda: defaultdict(set))
    teacher_daily_courses = defaultdict(lambda: defaultdict(list))

    for row in results:
        teacher_id = row["teacher_id"]
        weekday = row["week_day"]

        # 判断时段
        if row["start_slot"] <= 5:
            period = "morning"
        elif row["start_slot"] <= 10:
            period = "afternoon"
        else:
            period = "evening"

        teacher_daily_campus[teacher_id][(weekday, period)].add(row["campus_id"])
        teacher_daily_courses[teacher_id][(weekday, period)].append(
            {"course": row["course_name"], "campus": row["campus_name"]}
        )

    issues = []

    for teacher_id, daily_campus in teacher_daily_campus.items():
        for (weekday, period), campuses in daily_campus.items():
            if len(campuses) > 1:
                teacher_name = next(
                    (
                        r["teacher_name"]
                        for r in results
                        if r["teacher_id"] == teacher_id
                    ),
                    teacher_id,
                )

                courses = teacher_daily_courses[teacher_id][(weekday, period)]
                period_name = {
                    "morning": "上午",
                    "afternoon": "下午",
                    "evening": "晚上",
                }[period]

                issues.append(
                    {
                        "teacher_name": teacher_name,
                        "weekday": day_names[weekday],
                        "period": period_name,
                        "campus_count": len(campuses),
                        "campuses": list({c["campus"] for c in courses}),
                        "courses": [c["course"] for c in courses],
                    }
                )

    issues.sort(key=lambda x: x["campus_count"], reverse=True)

    if issues:
        print(f"\n⚠️  教师需跨校区: {len(issues)} 处")
        for i, issue in enumerate(issues[:10], 1):
            print(f"  {i}. {issue['teacher_name']}")
            print(
                f"     {issue['weekday']}{issue['period']} 跨{issue['campus_count']}个校区"
            )
            print(f"     校区: {', '.join(issue['campuses'])}")
            print(f"     课程: {', '.join(issue['courses'][:2])}")

        if len(issues) > 10:
            print(f"  ... 还有 {len(issues) - 10} 处未显示")
    else:
        print("✅ 无教师跨校区通勤问题")

    return issues


def generate_comprehensive_report(
    version_id,
    preference_violations,
    utilization_issues,
    student_overload,
    task_relation_violations,
    time_preference_issues,
    continuity_issues,
    commute_issues,
):
    """生成综合Excel报告"""
    print(f"\n{'='*70}")
    print("生成详细Excel报告...")
    print(f"{'='*70}\n")

    wb = Workbook()

    # 样式定义
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    warning_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    error_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )

    # 1. 教师偏好违反
    if preference_violations["avoided"]:
        ws = wb.create_sheet("教师偏好违反")
        ws.append(["序号", "教师", "课程", "实际时间", "避免时段", "教室", "惩罚分数"])

        for col in range(1, 8):
            ws.cell(1, col).fill = header_fill
            ws.cell(1, col).font = header_font

        for idx, v in enumerate(preference_violations["avoided"], 1):
            ws.append(
                [
                    idx,
                    v["teacher_name"],
                    v["course_name"],
                    f"{v['weekday']} {v['time']}",
                    v["avoid_time"],
                    v["classroom"],
                    v["penalty"],
                ]
            )

    # 2. 教室利用率问题
    if utilization_issues["waste"]:
        ws = wb.create_sheet("教室利用率低")
        ws.append(
            ["序号", "课程", "教室", "时间", "容量", "学生数", "利用率%", "浪费座位"]
        )

        for col in range(1, 9):
            ws.cell(1, col).fill = header_fill
            ws.cell(1, col).font = header_font

        for idx, issue in enumerate(utilization_issues["waste"], 1):
            ws.append(
                [
                    idx,
                    issue["course"],
                    issue["classroom"],
                    f"{issue['weekday']} {issue['time']}",
                    issue["capacity"],
                    issue["students"],
                    f"{issue['utilization']:.1f}",
                    issue["waste_seats"],
                ]
            )

    # 3. 学生负荷过重
    if student_overload:
        ws = wb.create_sheet("学生负荷过重")
        ws.append(["序号", "班级", "星期", "总节数", "课程数", "课程列表"])

        for col in range(1, 7):
            ws.cell(1, col).fill = header_fill
            ws.cell(1, col).font = header_font

        for idx, issue in enumerate(student_overload, 1):
            ws.append(
                [
                    idx,
                    issue["class_name"],
                    issue["weekday"],
                    issue["total_slots"],
                    issue["course_count"],
                    ", ".join(issue["courses"]),
                ]
            )

    # 4. 任务关系约束违反
    if task_relation_violations:
        ws = wb.create_sheet("任务关系约束违反")
        ws.append(
            ["序号", "课程", "约束类型", "问题描述", "第1次课", "第2次课", "惩罚分数"]
        )

        for col in range(1, 8):
            ws.cell(1, col).fill = header_fill
            ws.cell(1, col).font = header_font

        for idx, v in enumerate(task_relation_violations, 1):
            ws.append(
                [
                    idx,
                    v["course"],
                    v["constraint_type"],
                    v["reason"],
                    v["task_a_time"],
                    v["task_b_time"],
                    v["penalty"],
                ]
            )

    # 5. 时段偏好问题
    if time_preference_issues["night"] or time_preference_issues["weekend"]:
        ws = wb.create_sheet("课程时段问题")
        ws.append(["序号", "问题类型", "课程", "性质", "时间"])

        for col in range(1, 6):
            ws.cell(1, col).fill = header_fill
            ws.cell(1, col).font = header_font

        idx = 1
        for issue in time_preference_issues["night"]:
            ws.append(
                [
                    idx,
                    "必修课在晚上",
                    issue["course"],
                    issue["nature"],
                    f"{issue['weekday']} {issue['time']}",
                ]
            )
            idx += 1

        for issue in time_preference_issues["weekend"]:
            ws.append(
                [
                    idx,
                    "必修课在周末",
                    issue["course"],
                    issue["nature"],
                    f"{issue['weekday']} {issue['time']}",
                ]
            )
            idx += 1

    # 删除默认的Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"未满足约束详细报告_版本{version_id}_{timestamp}.xlsx"
    wb.save(filename)

    print(f"✅ 详细报告已导出: {filename}\n")

    # 显示总结
    print("\n【优化建议总结】")
    print("=" * 70)

    total_issues = (
        len(preference_violations["avoided"])
        + len(utilization_issues["waste"])
        + len(student_overload)
        + len(task_relation_violations)
        + len(time_preference_issues["night"])
        + len(time_preference_issues["weekend"])
        + len(commute_issues)
    )

    if total_issues == 0:
        print("\n🎉 恭喜！当前排课方案质量优秀，所有软约束均已满足！")
    else:
        print(f"\n发现 {total_issues} 处可优化点，建议：")
        print("\n1. 调整遗传算法参数：")
        print("   - 增加种群大小 (--population 300)")
        print("   - 增加迭代代数 (--generations 200)")
        print("   - 调整惩罚分数权重")

        print("\n2. 调整约束配置：")
        print("   - 降低部分软约束的惩罚分数")
        print("   - 放宽教师偏好时间范围")
        print("   - 增加可用教室数量")

        print("\n3. 手动微调：")
        print("   - 使用 optimize_conflicts 功能优化容量冲突")
        print("   - 针对特定课程手动调整时间/教室")

    print(f"\n{'='*70}\n")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python show_unsatisfied_constraints.py <version_id>")
        print("示例: python show_unsatisfied_constraints.py 1")
        sys.exit(1)

    version_id = int(sys.argv[1])
    analyze_unsatisfied_constraints(version_id)
