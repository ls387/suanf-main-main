#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
查看排课结果
"""
import os
import pymysql
from collections import defaultdict
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


def format_weeks(weeks_str, start_week=None, end_week=None, week_pattern=None):
    """将逗号分隔的周次字符串格式化为易读格式 (例如: '1-16周' 或 '1-8,10-16周')

    Args:
        weeks_str: 逗号分隔的周次字符串 (来自 offering_weeks 表)
        start_week: 起始周 (来自 course_offerings)
        end_week: 结束周 (来自 course_offerings)
        week_pattern: 周次模式 (来自 course_offerings)
    """
    # 如果有明确的周次数据,使用它
    if weeks_str:
        try:
            weeks = sorted([int(w) for w in weeks_str.split(",") if w.strip()])
            if weeks:
                ranges = []
                start = weeks[0]
                prev = weeks[0]

                for w in weeks[1:]:
                    if w == prev + 1:
                        prev = w
                    else:
                        if start == prev:
                            ranges.append(str(start))
                        else:
                            ranges.append(f"{start}-{prev}")
                        start = w
                        prev = w

                if start == prev:
                    ranges.append(str(start))
                else:
                    ranges.append(f"{start}-{prev}")

                return ",".join(ranges) + "周"
        except:
            pass

    # 如果没有周次数据,根据 week_pattern 生成
    if week_pattern and start_week and end_week:
        if week_pattern == "SINGLE":
            return f"{start_week}-{end_week}周(单)"
        elif week_pattern == "DOUBLE":
            return f"{start_week}-{end_week}周(双)"
        elif week_pattern in ("CONTINUOUS", "ALL_WEEKS"):
            return f"{start_week}-{end_week}周"
        elif week_pattern == "CUSTOM":
            return f"{start_week}-{end_week}周"

    return "全学期"


def view_schedule(version_id):
    """查看排课结果"""

    conn = pymysql.connect(
        host=os.getenv("DB_HOST") or "localhost",
        port=int(os.getenv("DB_PORT") or "3306"),
        user=os.getenv("DB_USER") or "root",
        password=os.getenv("DB_PASSWORD") or "123456",
        database=os.getenv("DB_NAME") or "paike",
        charset="utf8mb4",
    )

    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)

        # 获取版本信息
        cursor.execute(
            "SELECT * FROM schedule_versions WHERE version_id = %s", (version_id,)
        )
        version = cursor.fetchone()

        if not version:
            print(f"错误：找不到版本ID为 {version_id} 的排课版本")
            return

        print("=" * 100)
        print(f"排课结果查看 - {version['version_name']} ({version['semester']})")
        print(f"状态: {version['status']}")
        print("=" * 100)

        # 获取排课结果（包含教师和班级信息）
        query = """
        SELECT 
            sr.schedule_id,
            sr.week_day,
            sr.start_slot,
            sr.end_slot,
            sr.classroom_id,
            tt.task_id,
            tt.offering_id,
            tt.slots_count,
            c.course_name,
            co.course_nature,
            GROUP_CONCAT(DISTINCT t.teacher_name SEPARATOR ', ') AS teacher_name,
            GROUP_CONCAT(DISTINCT t.teacher_id SEPARATOR ', ') AS teacher_ids,
            GROUP_CONCAT(DISTINCT cl.class_name SEPARATOR ', ') AS class_names,
            GROUP_CONCAT(DISTINCT ow.week_number ORDER BY ow.week_number SEPARATOR ',') AS weeks,
            co.start_week,
            co.end_week,
            co.week_pattern,
            cr.classroom_name,
            cr.capacity,
            co.student_count_estimate
        FROM schedules sr
        JOIN teaching_tasks tt ON sr.task_id = tt.task_id
        JOIN course_offerings co ON tt.offering_id = co.offering_id
        JOIN courses c ON co.course_id = c.course_id
        LEFT JOIN offering_teachers ot ON co.offering_id = ot.offering_id
        LEFT JOIN teachers t ON ot.teacher_id = t.teacher_id
        LEFT JOIN offering_classes oc ON co.offering_id = oc.offering_id
        LEFT JOIN classes cl ON oc.class_id = cl.class_id
        LEFT JOIN offering_weeks ow ON co.offering_id = ow.offering_id
        JOIN classrooms cr ON sr.classroom_id = cr.classroom_id
        WHERE sr.version_id = %s
        GROUP BY sr.schedule_id, sr.week_day, sr.start_slot, sr.end_slot, 
                 sr.classroom_id, tt.task_id, tt.offering_id, tt.slots_count, c.course_name, 
                 co.course_nature, co.start_week, co.end_week, co.week_pattern,
                 cr.classroom_name, cr.capacity, co.student_count_estimate
        ORDER BY sr.week_day, sr.start_slot
        """

        cursor.execute(query, (version_id,))
        results = cursor.fetchall()

        if not results:
            print("该版本还没有排课结果")
            return

        print(f"\n共有 {len(results)} 条排课记录\n")

        # 按星期分组显示
        day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        schedule_by_day = defaultdict(list)

        for result in results:
            schedule_by_day[result["week_day"]].append(result)

        # 显示每天的课表
        for day in range(1, 8):
            if day not in schedule_by_day:
                continue

            print(f"\n{'='*100}")
            print(f"{day_names[day]}")
            print(f"{'='*100}")
            print(
                f"{'节次':<6} {'课程名称':<25} {'性质':<8} {'教师':<12} {'教室':<15} {'人数/容量':<12} {'周次':<15}"
            )
            print("-" * 115)

            for item in schedule_by_day[day]:
                slots_str = f"{item['start_slot']}-{item['end_slot']}"

                student_count = item["student_count_estimate"] or 0
                capacity = item["capacity"] or 0
                utilization = f"{student_count}/{capacity}"

                teacher_name = item["teacher_name"] or "未分配"
                weeks_display = format_weeks(
                    item.get("weeks"),
                    item.get("start_week"),
                    item.get("end_week"),
                    item.get("week_pattern"),
                )

                print(
                    f"{slots_str:<6} {item['course_name']:<25} {item['course_nature']:<8} "
                    f"{teacher_name:<12} {item['classroom_name']:<15} {utilization:<12} {weeks_display:<15}"
                )

        # 统计信息
        print(f"\n{'='*100}")
        print("统计信息")
        print(f"{'='*100}")

        # 按课程性质统计
        nature_count = defaultdict(int)
        for result in results:
            nature_count[result["course_nature"]] += 1

        print("\n课程性质分布:")
        for nature, count in nature_count.items():
            print(f"  {nature}: {count} 门")

        # 按星期统计
        print("\n每日课程数量:")
        for day in range(1, 8):
            if day in schedule_by_day:
                print(f"  {day_names[day]}: {len(schedule_by_day[day])} 门课程")

        # 教室利用率
        total_utilization = 0
        count = 0
        for result in results:
            if (
                result["capacity"]
                and result["capacity"] > 0
                and result["student_count_estimate"]
            ):
                util = result["student_count_estimate"] / result["capacity"]
                total_utilization += util
                count += 1

        if count > 0:
            avg_utilization = total_utilization / count
            print(f"\n平均教室利用率: {avg_utilization:.1%}")

        print("\n" + "=" * 100)

        # 导出到 Excel（包含四个视角）
        export_to_excel(version, results, schedule_by_day, cursor)

    finally:
        cursor.close()
        conn.close()


def export_to_excel(version, results, schedule_by_day, cursor):
    """导出排课结果到 Excel（四个视角）"""

    # 创建工作簿
    wb = Workbook()

    # 删除默认工作表
    wb.remove(wb.active)

    # 样式定义
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    # ============ 视角1：按周期查看（原有的按天分组） ============
    for day in range(1, 8):
        if day not in schedule_by_day:
            continue

        ws = wb.create_sheet(title=f"{day_names[day]}")

        # 设置列宽
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 15

        # 写入标题
        headers = [
            "节次",
            "课程名称",
            "性质",
            "教师",
            "教室",
            "班级",
            "教室容量",
            "上课人数",
            "周次",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 写入数据
        for row_idx, item in enumerate(schedule_by_day[day], 2):
            slots_str = f"{item['start_slot']}-{item['end_slot']}"
            student_count = item["student_count_estimate"] or 0
            capacity = item["capacity"] or 0
            teacher_name = item["teacher_name"] or "未分配"
            class_names = item["class_names"] or "未分配"

            weeks_display = format_weeks(
                item.get("weeks"),
                item.get("start_week"),
                item.get("end_week"),
                item.get("week_pattern"),
            )

            data = [
                slots_str,
                item["course_name"],
                item["course_nature"],
                teacher_name,
                item["classroom_name"],
                class_names,
                capacity,
                student_count,
                weeks_display,
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # ============ 视角2：教师视角 ============
    # 获取所有教师及其课程
    cursor.execute(
        """
        SELECT DISTINCT t.teacher_id, t.teacher_name
        FROM teachers t
        JOIN offering_teachers ot ON t.teacher_id = ot.teacher_id
        JOIN teaching_tasks tt ON ot.offering_id = tt.offering_id
        JOIN schedules sr ON tt.task_id = sr.task_id
        WHERE sr.version_id = %s
        ORDER BY t.teacher_name
    """,
        (version["version_id"],),
    )
    teachers = cursor.fetchall()

    for teacher in teachers:
        teacher_id = str(teacher["teacher_id"])
        teacher_name = teacher["teacher_name"]

        # 获取该教师的所有课程（检查 teacher_ids 字段）
        teacher_schedule = [
            r
            for r in results
            if r.get("teacher_ids") and teacher_id in r["teacher_ids"].split(", ")
        ]

        if not teacher_schedule:
            continue

        # 按周几分组
        teacher_by_day = defaultdict(list)
        for item in teacher_schedule:
            teacher_by_day[item["week_day"]].append(item)

        ws = wb.create_sheet(title=f"教师_{teacher_name[:10]}")

        # 设置列宽
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 15

        # 写入标题
        headers = [
            "星期",
            "节次",
            "课程名称",
            "教室",
            "班级",
            "教室容量",
            "上课人数",
            "周次",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 写入数据
        row_idx = 2
        for day in range(1, 8):
            if day in teacher_by_day:
                for item in teacher_by_day[day]:
                    slots_str = f"{item['start_slot']}-{item['end_slot']}"
                    student_count = item["student_count_estimate"] or 0
                    capacity = item["capacity"] or 0
                    class_names = item["class_names"] or "未分配"
                    weeks_display = format_weeks(
                        item.get("weeks"),
                        item.get("start_week"),
                        item.get("end_week"),
                        item.get("week_pattern"),
                    )

                    data = [
                        day_names[day],
                        slots_str,
                        item["course_name"],
                        item["classroom_name"],
                        class_names,
                        capacity,
                        student_count,
                        weeks_display,
                    ]

                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row_idx, column=col, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    row_idx += 1

    # ============ 视角3：班级视角 ============
    # 获取所有班级及其课程
    cursor.execute(
        """
        SELECT DISTINCT cl.class_id, cl.class_name
        FROM classes cl
        JOIN offering_classes oc ON cl.class_id = oc.class_id
        JOIN course_offerings co ON oc.offering_id = co.offering_id
        JOIN teaching_tasks tt ON co.offering_id = tt.offering_id
        JOIN schedules sr ON tt.task_id = sr.task_id
        WHERE sr.version_id = %s
        ORDER BY cl.class_name
    """,
        (version["version_id"],),
    )
    classes = cursor.fetchall()

    for class_info in classes:
        class_id = class_info["class_id"]
        class_name = class_info["class_name"]

        # 获取该班级的所有课程
        class_schedule = []
        for r in results:
            if (
                r["class_names"]
                and str(class_id) in r["class_names"]
                or class_name in (r["class_names"] or "")
            ):
                class_schedule.append(r)

        if not class_schedule:
            continue

        # 按周几分组
        class_by_day = defaultdict(list)
        for item in class_schedule:
            class_by_day[item["week_day"]].append(item)

        ws = wb.create_sheet(title=f"班级_{class_name[:10]}")

        # 设置列宽
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 15

        # 写入标题
        headers = ["星期", "节次", "课程名称", "教室", "教师", "性质", "周次"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 写入数据
        row_idx = 2
        for day in range(1, 8):
            if day in class_by_day:
                for item in class_by_day[day]:
                    slots_str = f"{item['start_slot']}-{item['end_slot']}"
                    teacher_name = item["teacher_name"] or "未分配"
                    weeks_display = format_weeks(
                        item.get("weeks"),
                        item.get("start_week"),
                        item.get("end_week"),
                        item.get("week_pattern"),
                    )

                    data = [
                        day_names[day],
                        slots_str,
                        item["course_name"],
                        item["classroom_name"],
                        teacher_name,
                        item["course_nature"],
                        weeks_display,
                    ]

                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row_idx, column=col, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    row_idx += 1

    # ============ 视角4：教室视角 ============
    # 获取所有使用的教室
    cursor.execute(
        """
        SELECT DISTINCT cr.classroom_id, cr.classroom_name, cr.capacity
        FROM classrooms cr
        JOIN schedules sr ON cr.classroom_id = sr.classroom_id
        WHERE sr.version_id = %s
        ORDER BY cr.classroom_name
    """,
        (version["version_id"],),
    )
    classrooms = cursor.fetchall()

    for classroom in classrooms:
        classroom_id = classroom["classroom_id"]
        classroom_name = classroom["classroom_name"]
        classroom_capacity = classroom["capacity"]

        # 获取该教室的所有课程
        classroom_schedule = [r for r in results if r["classroom_id"] == classroom_id]

        if not classroom_schedule:
            continue

        # 按周几分组
        classroom_by_day = defaultdict(list)
        for item in classroom_schedule:
            classroom_by_day[item["week_day"]].append(item)

        ws = wb.create_sheet(title=f"教室_{classroom_name[:10]}")

        # 设置列宽
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 15

        # 写入标题
        headers = [
            "星期",
            "节次",
            "课程名称",
            "班级",
            "教师",
            "教室容量",
            "上课人数",
            "周次",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 写入数据
        row_idx = 2
        for day in range(1, 8):
            if day in classroom_by_day:
                for item in classroom_by_day[day]:
                    slots_str = f"{item['start_slot']}-{item['end_slot']}"
                    student_count = item["student_count_estimate"] or 0
                    teacher_name = item["teacher_name"] or "未分配"
                    class_names = item["class_names"] or "未分配"
                    weeks_display = format_weeks(
                        item.get("weeks"),
                        item.get("start_week"),
                        item.get("end_week"),
                        item.get("week_pattern"),
                    )

                    data = [
                        day_names[day],
                        slots_str,
                        item["course_name"],
                        class_names,
                        teacher_name,
                        classroom_capacity,
                        student_count,
                        weeks_display,
                    ]

                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row_idx, column=col, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    row_idx += 1

    # 创建汇总表
    summary_ws = wb.create_sheet(title="汇总", index=0)
    summary_ws.column_dimensions["A"].width = 20
    summary_ws.column_dimensions["B"].width = 30

    # 写入版本信息
    summary_ws["A1"] = "排课版本"
    summary_ws["B1"] = version["version_name"]
    summary_ws["A2"] = "学期"
    summary_ws["B2"] = version["semester"]
    summary_ws["A3"] = "状态"
    summary_ws["B3"] = version["status"]
    summary_ws["A4"] = "总课程数"
    summary_ws["B4"] = len(results)
    summary_ws["A5"] = "导出时间"
    summary_ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 应用样式到汇总表
    for row in range(1, 6):
        summary_ws.cell(row=row, column=1).font = Font(bold=True)

    # 课程性质统计
    summary_ws["A7"] = "课程性质分布"
    summary_ws["A7"].font = Font(bold=True, size=12)

    nature_count = defaultdict(int)
    for result in results:
        nature_count[result["course_nature"]] += 1

    row = 8
    for nature, count in nature_count.items():
        summary_ws.cell(row=row, column=1, value=nature)
        summary_ws.cell(row=row, column=2, value=f"{count} 门")
        row += 1

    # 每日课程数统计
    summary_ws.cell(row=row + 1, column=1, value="每日课程数").font = Font(
        bold=True, size=12
    )
    row += 2
    for day in range(1, 8):
        if day in schedule_by_day:
            summary_ws.cell(row=row, column=1, value=day_names[day])
            summary_ws.cell(
                row=row, column=2, value=f"{len(schedule_by_day[day])} 门课程"
            )
            row += 1

    # 保存文件
    filename = f"排课结果_版本{version['version_id']}_{version['semester']}.xlsx"
    wb.save(filename)

    print(f"\n√ 排课结果已导出到: {filename}")
    print(f"  文件位置: {os.path.abspath(filename)}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python view_schedule.py <version_id>")
        print("示例: python view_schedule.py 1")
        sys.exit(1)

    version_id = int(sys.argv[1])
    view_schedule(version_id)
