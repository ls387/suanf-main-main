#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
分析排课结果中的冲突问题
支持导出冲突Excel和优化冲突
"""
import os
import sys
import pymysql
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 设置标准输出编码为UTF-8
if sys.platform == "win32":
    import io

    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


def has_week_overlap(weeks1, weeks2):
    """检查两个周次集合是否有重叠

    Args:
        weeks1: 第一个课程的周次集合
        weeks2: 第二个课程的周次集合

    Returns:
        bool: 如果有重叠返回True,否则返回False
    """
    if not weeks1 or not weeks2:
        # 如果任一课程没有周次信息,保守起见认为有冲突
        return True

    return len(weeks1 & weeks2) > 0


def export_conflicts_to_excel(
    class_conflicts,
    teacher_conflicts,
    classroom_conflicts,
    capacity_violations,
    version_id,
):
    """导出冲突到Excel"""
    wb = Workbook()

    # 样式定义
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    conflict_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )

    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    # 班级冲突表
    if wb.active:
        ws1 = wb.active
        ws1.title = "班级冲突"
    else:
        ws1 = wb.create_sheet("班级冲突")

    ws1.append(
        [
            "冲突序号",
            "班级",
            "星期",
            "节次",
            "课程1",
            "教师1",
            "教室1",
            "课程2",
            "教师2",
            "教室2",
        ]
    )
    for col in range(1, 11):
        ws1.cell(1, col).fill = header_fill
        ws1.cell(1, col).font = header_font
        ws1.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

    for idx, conflict in enumerate(class_conflicts, 1):
        items = conflict["items"]
        # 格式化时间段
        time_range = (
            f"{conflict['overlap_start']}-{conflict['overlap_end']}节"
            if conflict["overlap_start"] != conflict["overlap_end"]
            else f"{conflict['overlap_start']}节"
        )
        ws1.append(
            [
                idx,
                conflict["class_name"],
                day_names[conflict["weekday"]],
                time_range,
                items[0]["course"],
                items[0]["teacher"],
                items[0]["classroom"],
                items[1]["course"] if len(items) > 1 else "",
                items[1]["teacher"] if len(items) > 1 else "",
                items[1]["classroom"] if len(items) > 1 else "",
            ]
        )
        for col in range(1, 11):
            ws1.cell(idx + 1, col).fill = conflict_fill

    # 调整列宽
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 8
    ws1.column_dimensions["D"].width = 8
    for col in ["E", "F", "G", "H", "I", "J"]:
        ws1.column_dimensions[col].width = 20

    # 教师冲突表
    ws2 = wb.create_sheet("教师冲突")
    ws2.append(["冲突序号", "教师", "星期", "节次", "课程1", "教室1", "课程2", "教室2"])
    for col in range(1, 9):
        ws2.cell(1, col).fill = header_fill
        ws2.cell(1, col).font = header_font
        ws2.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

    for idx, conflict in enumerate(teacher_conflicts, 1):
        items = conflict["items"]
        # 格式化时间段
        time_range = (
            f"{conflict['overlap_start']}-{conflict['overlap_end']}节"
            if conflict["overlap_start"] != conflict["overlap_end"]
            else f"{conflict['overlap_start']}节"
        )
        ws2.append(
            [
                idx,
                conflict["teacher"],
                day_names[conflict["weekday"]],
                time_range,
                items[0]["course"],
                items[0]["classroom"],
                items[1]["course"] if len(items) > 1 else "",
                items[1]["classroom"] if len(items) > 1 else "",
            ]
        )
        for col in range(1, 9):
            ws2.cell(idx + 1, col).fill = conflict_fill

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 8
    ws2.column_dimensions["D"].width = 8
    for col in ["E", "F", "G", "H"]:
        ws2.column_dimensions[col].width = 20

    # 教室冲突表
    ws3 = wb.create_sheet("教室冲突")
    ws3.append(["冲突序号", "教室", "星期", "节次", "课程1", "教师1", "课程2", "教师2"])
    for col in range(1, 9):
        ws3.cell(1, col).fill = header_fill
        ws3.cell(1, col).font = header_font
        ws3.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

    for idx, conflict in enumerate(classroom_conflicts, 1):
        items = conflict["items"]
        # 格式化时间段
        time_range = (
            f"{conflict['overlap_start']}-{conflict['overlap_end']}节"
            if conflict["overlap_start"] != conflict["overlap_end"]
            else f"{conflict['overlap_start']}节"
        )
        ws3.append(
            [
                idx,
                conflict["classroom"],
                day_names[conflict["weekday"]],
                time_range,
                items[0]["course"],
                items[0]["teacher"],
                items[1]["course"] if len(items) > 1 else "",
                items[1]["teacher"] if len(items) > 1 else "",
            ]
        )
        for col in range(1, 9):
            ws3.cell(idx + 1, col).fill = conflict_fill

    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 8
    ws3.column_dimensions["D"].width = 8
    for col in ["E", "F", "G", "H"]:
        ws3.column_dimensions[col].width = 20

    # 容量不足冲突表
    ws4 = wb.create_sheet("容量不足冲突")
    ws4.append(
        ["序号", "课程", "星期", "节次", "教室", "教室容量", "学生数", "缺少座位"]
    )
    for col in range(1, 9):
        ws4.cell(1, col).fill = header_fill
        ws4.cell(1, col).font = header_font
        ws4.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

    for idx, item in enumerate(capacity_violations, 1):
        # 格式化时间段
        time_range = (
            f"{item['start_slot']}-{item['end_slot']}节"
            if item["start_slot"] != item["end_slot"]
            else f"{item['start_slot']}节"
        )
        ws4.append(
            [
                idx,
                item["course"],
                day_names[item["weekday"]],
                time_range,
                item["classroom"],
                item["capacity"],
                item["students"],
                item["shortage"],
            ]
        )
        for col in range(1, 9):
            ws4.cell(idx + 1, col).fill = conflict_fill

    ws4.column_dimensions["A"].width = 8
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 8
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 15
    ws4.column_dimensions["F"].width = 10
    ws4.column_dimensions["G"].width = 10
    ws4.column_dimensions["H"].width = 12

    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"排课冲突报告_版本{version_id}_{timestamp}.xlsx"
    wb.save(filename)
    print(f"\n✓ 冲突报告已导出到: {filename}")
    return filename


def analyze_schedule_conflicts(version_id):
    """分析指定版本的排课冲突"""

    # 连接数据库
    conn = pymysql.connect(
        host=os.getenv("DB_HOST") or "localhost",
        port=int(os.getenv("DB_PORT") or "3306"),
        user=os.getenv("DB_USER") or "pk",
        password=os.getenv("DB_PASSWORD") or "123456",
        database=os.getenv("DB_NAME") or "paike",
        charset="utf8mb4",
    )

    try:
        cursor = conn.cursor(pymysql.cursors.DictCursor)

        # 获取排课结果
        query = """
        SELECT 
            sr.schedule_id,
            sr.task_id,
            sr.classroom_id,
            sr.week_day,
            sr.start_slot,
            tt.slots_count,
            tt.offering_id,
            c.course_name,
            cr.classroom_name,
            co.start_week,
            co.end_week,
            co.week_pattern,
            GROUP_CONCAT(DISTINCT t.teacher_name SEPARATOR ', ') AS teacher_name,
            GROUP_CONCAT(DISTINCT t.teacher_id SEPARATOR ', ') AS teacher_ids
        FROM schedules sr
        JOIN teaching_tasks tt ON sr.task_id = tt.task_id
        JOIN course_offerings co ON tt.offering_id = co.offering_id
        JOIN courses c ON co.course_id = c.course_id
        JOIN classrooms cr ON sr.classroom_id = cr.classroom_id
        LEFT JOIN offering_teachers ot ON co.offering_id = ot.offering_id
        LEFT JOIN teachers t ON ot.teacher_id = t.teacher_id
        WHERE sr.version_id = %s
        GROUP BY sr.schedule_id, sr.task_id, sr.classroom_id, sr.week_day, 
                 sr.start_slot, tt.slots_count, tt.offering_id, c.course_name, cr.classroom_name,
                 co.start_week, co.end_week, co.week_pattern
        ORDER BY sr.week_day, sr.start_slot
        """

        cursor.execute(query, (version_id,))
        results = cursor.fetchall()

        # 加载周次信息
        offering_weeks = {}
        weeks_query = "SELECT offering_id, week_number FROM offering_weeks"
        cursor.execute(weeks_query)
        for row in cursor.fetchall():
            offering_id = row["offering_id"]
            week_number = row["week_number"]
            if offering_id not in offering_weeks:
                offering_weeks[offering_id] = set()
            offering_weeks[offering_id].add(week_number)

        # 为每个结果生成周次集合
        def get_weeks(result):
            """根据offering_id获取周次集合"""
            offering_id = result["offering_id"]
            if offering_id in offering_weeks:
                return offering_weeks[offering_id]

            # 如果没有自定义周次,根据week_pattern生成
            start_week = result.get("start_week") or 1
            end_week = result.get("end_week") or 18
            week_pattern = result.get("week_pattern") or "CONTINUOUS"

            if week_pattern == "CONTINUOUS":
                return set(range(start_week, end_week + 1))
            elif week_pattern == "SINGLE":
                return set(w for w in range(start_week, end_week + 1) if w % 2 == 1)
            elif week_pattern == "DOUBLE":
                return set(w for w in range(start_week, end_week + 1) if w % 2 == 0)
            else:
                return set(range(start_week, end_week + 1))

        # 获取任务-班级关系
        task_classes = defaultdict(list)
        for result in results:
            task_id = result["task_id"]
            class_query = """
            SELECT cl.class_id, cl.class_name
            FROM offering_classes oc
            JOIN classes cl ON oc.class_id = cl.class_id
            JOIN teaching_tasks tt ON oc.offering_id = tt.offering_id
            WHERE tt.task_id = %s
            """
            cursor.execute(class_query, (task_id,))
            classes = cursor.fetchall()
            task_classes[task_id] = classes

        print("=" * 80)
        print(f"排课版本 {version_id} 冲突分析报告")
        print("=" * 80)

        # 分析容量冲突和浪费
        print("\n【教室容量分析】")
        capacity_violations = []  # 容量不足
        capacity_conflicts_data = []  # 容量冲突数据（用于优化）
        high_waste = []  # 容量浪费严重
        high_waste_data = []  # 容量浪费数据（用于优化）
        total_waste_seats = 0

        # 获取学生人数信息
        for result in results:
            # 查询该任务的学生总数
            student_query = """
            SELECT SUM(cl.student_count) as total_students
            FROM offering_classes oc
            JOIN classes cl ON oc.class_id = cl.class_id
            JOIN teaching_tasks tt ON oc.offering_id = tt.offering_id
            WHERE tt.task_id = %s
            """
            cursor.execute(student_query, (result["task_id"],))
            student_result = cursor.fetchone()
            student_count = int(student_result["total_students"] or 0)  # 转换为整数

            # 查询教室容量
            classroom_query = "SELECT capacity FROM classrooms WHERE classroom_id = %s"
            cursor.execute(classroom_query, (result["classroom_id"],))
            classroom_result = cursor.fetchone()
            classroom_capacity = (
                int(classroom_result["capacity"]) if classroom_result else 0
            )  # 转换为整数

            # 检查容量冲突
            if classroom_capacity < student_count:
                end_slot = result["start_slot"] + result["slots_count"] - 1
                capacity_violations.append(
                    {
                        "course": result["course_name"],
                        "classroom": result["classroom_name"],
                        "capacity": classroom_capacity,
                        "students": student_count,
                        "shortage": student_count - classroom_capacity,
                        "weekday": result["week_day"],
                        "start_slot": result["start_slot"],
                        "end_slot": end_slot,
                    }
                )
                # 记录详细数据用于优化
                capacity_conflicts_data.append(
                    {
                        "schedule_id": result["schedule_id"],
                        "task_id": result["task_id"],
                        "course": result["course_name"],
                        "classroom_id": result["classroom_id"],
                        "classroom": result["classroom_name"],
                        "capacity": classroom_capacity,
                        "students": student_count,
                        "shortage": student_count - classroom_capacity,
                        "weekday": result["week_day"],
                        "start_slot": result["start_slot"],
                        "slots_count": result["slots_count"],
                    }
                )

            # 检查容量浪费
            if classroom_capacity > 0 and student_count > 0:
                waste_seats = classroom_capacity - student_count
                total_waste_seats += waste_seats
                waste_ratio = waste_seats / classroom_capacity
                utilization = student_count / classroom_capacity

                # 根据班级规模判断是否浪费过大
                max_waste_ratio = (
                    0.5 if student_count < 30 else (0.4 if student_count < 60 else 0.3)
                )
                if waste_ratio > max_waste_ratio:
                    high_waste.append(
                        {
                            "course": result["course_name"],
                            "classroom": result["classroom_name"],
                            "capacity": classroom_capacity,
                            "students": student_count,
                            "waste_seats": waste_seats,
                            "waste_ratio": waste_ratio,
                            "utilization": utilization,
                        }
                    )
                    # 记录详细数据用于优化
                    high_waste_data.append(
                        {
                            "schedule_id": result["schedule_id"],
                            "task_id": result["task_id"],
                            "course": result["course_name"],
                            "classroom_id": result["classroom_id"],
                            "classroom": result["classroom_name"],
                            "capacity": classroom_capacity,
                            "students": student_count,
                            "waste_seats": waste_seats,
                            "waste_ratio": waste_ratio,
                            "weekday": result["week_day"],
                            "start_slot": result["start_slot"],
                            "slots_count": result["slots_count"],
                        }
                    )

        # 报告容量冲突
        if capacity_violations:
            print(f"\n容量不足冲突: {len(capacity_violations)} 个")
            for i, item in enumerate(capacity_violations[:10], 1):
                print(
                    f"  [{i}] {item['course']}: 教室 {item['classroom']} "
                    f"(容量{item['capacity']}) < 学生数{item['students']}, "
                    f"缺少 {item['shortage']} 个座位"
                )
            if len(capacity_violations) > 10:
                print(f"  ... 还有 {len(capacity_violations) - 10} 个容量不足问题")
        else:
            print("✓ 无容量不足问题")

        # 报告容量浪费
        print(f"\n总计浪费座位数: {total_waste_seats} 个")
        if len(results) > 0:
            print(f"平均每节课浪费: {total_waste_seats / len(results):.1f} 个座位")

        if high_waste:
            print(f"\n容量浪费严重课程: {len(high_waste)} 个")
            # 按浪费率排序
            high_waste.sort(key=lambda x: x["waste_ratio"], reverse=True)
            for i, item in enumerate(high_waste[:10], 1):
                print(
                    f"  [{i}] {item['course']}: 教室 {item['classroom']} "
                    f"(容量{item['capacity']}) > 学生数{item['students']}, "
                    f"浪费 {item['waste_seats']} 个座位 (浪费率{item['waste_ratio']:.1%}, 利用率{item['utilization']:.1%})"
                )
            if len(high_waste) > 10:
                print(f"  ... 还有 {len(high_waste) - 10} 个浪费严重的课程")
        else:
            print("✓ 无严重容量浪费")

        # 分析班级冲突
        print("\n【班级时间冲突分析】")
        class_schedule = defaultdict(list)
        class_conflicts_data = []  # 用于导出

        for result in results:
            task_id = result["task_id"]
            start_slot = result["start_slot"]
            end_slot = start_slot + result["slots_count"] - 1
            week_day = result["week_day"]
            weeks = get_weeks(result)

            for cls in task_classes[task_id]:
                class_id = cls["class_id"]
                for slot in range(start_slot, end_slot + 1):
                    time_key = (week_day, slot)
                    class_schedule[class_id].append(
                        {
                            "time": time_key,
                            "course": result["course_name"],
                            "teacher": result["teacher_name"],
                            "classroom": result["classroom_name"],
                            "class_name": cls["class_name"],
                            "weeks": weeks,
                        }
                    )

        # 检测冲突(考虑周次重叠) - 按课程对去重
        conflicts_found = 0
        conflict_pairs = set()  # 用于去重：(class_id, schedule_id1, schedule_id2)

        # 重新构建schedule，包含schedule_id
        class_schedule_with_id = defaultdict(list)
        for result in results:
            task_id = result["task_id"]
            schedule_id = result["schedule_id"]
            start_slot = result["start_slot"]
            end_slot = start_slot + result["slots_count"] - 1
            week_day = result["week_day"]
            weeks = get_weeks(result)

            for cls in task_classes[task_id]:
                class_id = cls["class_id"]
                class_schedule_with_id[class_id].append(
                    {
                        "schedule_id": schedule_id,
                        "weekday": week_day,
                        "start_slot": start_slot,
                        "end_slot": end_slot,
                        "course": result["course_name"],
                        "teacher": result["teacher_name"],
                        "classroom": result["classroom_name"],
                        "class_name": cls["class_name"],
                        "weeks": weeks,
                    }
                )

        # 检测冲突
        for class_id, schedules in class_schedule_with_id.items():
            # 检查每对课程是否有时间和周次重叠
            for i in range(len(schedules)):
                for j in range(i + 1, len(schedules)):
                    s1, s2 = schedules[i], schedules[j]

                    # 检查时间是否重叠
                    if s1["weekday"] != s2["weekday"]:
                        continue

                    # 检查时间段是否重叠
                    if (
                        s1["end_slot"] < s2["start_slot"]
                        or s2["end_slot"] < s1["start_slot"]
                    ):
                        continue

                    # 检查周次是否重叠
                    if not has_week_overlap(s1["weeks"], s2["weeks"]):
                        continue

                    # 确认是真冲突，检查是否已记录
                    pair_key = tuple(sorted([s1["schedule_id"], s2["schedule_id"]]))
                    if (class_id, pair_key) in conflict_pairs:
                        continue

                    conflict_pairs.add((class_id, pair_key))
                    conflicts_found += 1

                    day_names = [
                        "",
                        "周一",
                        "周二",
                        "周三",
                        "周四",
                        "周五",
                        "周六",
                        "周日",
                    ]
                    overlap_start = max(s1["start_slot"], s2["start_slot"])
                    overlap_end = min(s1["end_slot"], s2["end_slot"])

                    print(f"\n冲突 #{conflicts_found}:")
                    print(f"  班级: {s1['class_name']}")
                    print(
                        f"  时间: {day_names[s1['weekday']]} 第{overlap_start}-{overlap_end}节"
                    )
                    print(
                        f"  课程1: {s1['course']} ({s1['start_slot']}-{s1['end_slot']}节) - {s1['teacher']} - {s1['classroom']}"
                    )
                    print(
                        f"  课程2: {s2['course']} ({s2['start_slot']}-{s2['end_slot']}节) - {s2['teacher']} - {s2['classroom']}"
                    )

                    # 记录冲突数据用于导出
                    class_conflicts_data.append(
                        {
                            "class_name": s1["class_name"],
                            "weekday": s1["weekday"],
                            "overlap_start": overlap_start,
                            "overlap_end": overlap_end,
                            "items": [s1, s2],
                        }
                    )

        if conflicts_found == 0:
            print("✓ 未发现班级冲突")
        else:
            print(f"\n⚠ 共发现 {conflicts_found} 处班级冲突")

        # 分析教师冲突
        print("\n【教师时间冲突分析】")
        teacher_schedule = defaultdict(list)
        teacher_conflicts_data = []  # 用于导出

        for result in results:
            # 从 teacher_ids 字段获取教师ID列表
            teacher_ids_str = result.get("teacher_ids", "")
            if not teacher_ids_str:
                continue

            teacher_ids = teacher_ids_str.split(", ")
            start_slot = result["start_slot"]
            end_slot = start_slot + result["slots_count"] - 1
            week_day = result["week_day"]
            weeks = get_weeks(result)

            for teacher_id in teacher_ids:
                for slot in range(start_slot, end_slot + 1):
                    time_key = (week_day, slot)
                    teacher_schedule[teacher_id].append(
                        {
                            "time": time_key,
                            "course": result["course_name"],
                            "teacher": result["teacher_name"],
                            "classroom": result["classroom_name"],
                            "weeks": weeks,
                        }
                    )

        # 重新构建教师schedule，包含schedule_id
        teacher_schedule_with_id = defaultdict(list)
        for result in results:
            teacher_ids_str = result.get("teacher_ids", "")
            if not teacher_ids_str:
                continue

            schedule_id = result["schedule_id"]
            teacher_ids = teacher_ids_str.split(", ")
            start_slot = result["start_slot"]
            end_slot = start_slot + result["slots_count"] - 1
            week_day = result["week_day"]
            weeks = get_weeks(result)

            for teacher_id in teacher_ids:
                teacher_schedule_with_id[teacher_id].append(
                    {
                        "schedule_id": schedule_id,
                        "weekday": week_day,
                        "start_slot": start_slot,
                        "end_slot": end_slot,
                        "course": result["course_name"],
                        "teacher": result["teacher_name"],
                        "classroom": result["classroom_name"],
                        "weeks": weeks,
                    }
                )

        teacher_conflicts = 0
        teacher_conflict_pairs = set()  # 用于去重

        for teacher_id, schedules in teacher_schedule_with_id.items():
            # 检查每对课程是否有时间和周次重叠
            for i in range(len(schedules)):
                for j in range(i + 1, len(schedules)):
                    s1, s2 = schedules[i], schedules[j]

                    # 检查时间是否重叠
                    if s1["weekday"] != s2["weekday"]:
                        continue

                    # 检查时间段是否重叠
                    if (
                        s1["end_slot"] < s2["start_slot"]
                        or s2["end_slot"] < s1["start_slot"]
                    ):
                        continue

                    # 检查周次是否重叠
                    if not has_week_overlap(s1["weeks"], s2["weeks"]):
                        continue

                    # 确认是真冲突，检查是否已记录
                    pair_key = tuple(sorted([s1["schedule_id"], s2["schedule_id"]]))
                    if (teacher_id, pair_key) in teacher_conflict_pairs:
                        continue

                    teacher_conflict_pairs.add((teacher_id, pair_key))
                    teacher_conflicts += 1

                    day_names = [
                        "",
                        "周一",
                        "周二",
                        "周三",
                        "周四",
                        "周五",
                        "周六",
                        "周日",
                    ]
                    overlap_start = max(s1["start_slot"], s2["start_slot"])
                    overlap_end = min(s1["end_slot"], s2["end_slot"])

                    print(f"\n冲突 #{teacher_conflicts}:")
                    print(f"  教师: {s1['teacher']}")
                    print(
                        f"  时间: {day_names[s1['weekday']]} 第{overlap_start}-{overlap_end}节"
                    )
                    print(
                        f"  课程1: {s1['course']} ({s1['start_slot']}-{s1['end_slot']}节) - {s1['classroom']}"
                    )
                    print(
                        f"  课程2: {s2['course']} ({s2['start_slot']}-{s2['end_slot']}节) - {s2['classroom']}"
                    )

                    # 记录冲突数据用于导出
                    teacher_conflicts_data.append(
                        {
                            "teacher": s1["teacher"],
                            "weekday": s1["weekday"],
                            "overlap_start": overlap_start,
                            "overlap_end": overlap_end,
                            "items": [s1, s2],
                        }
                    )

        if teacher_conflicts == 0:
            print("✓ 未发现教师冲突")
        else:
            print(f"\n⚠ 共发现 {teacher_conflicts} 处教师冲突")

        # 分析教室冲突
        print("\n【教室时间冲突分析】")
        classroom_schedule_with_id = defaultdict(list)
        classroom_conflicts_data = []  # 用于导出

        for result in results:
            classroom_id = result["classroom_id"]
            schedule_id = result["schedule_id"]
            start_slot = result["start_slot"]
            end_slot = start_slot + result["slots_count"] - 1
            week_day = result["week_day"]
            weeks = get_weeks(result)

            classroom_schedule_with_id[classroom_id].append(
                {
                    "schedule_id": schedule_id,
                    "weekday": week_day,
                    "start_slot": start_slot,
                    "end_slot": end_slot,
                    "course": result["course_name"],
                    "teacher": result["teacher_name"],
                    "classroom": result["classroom_name"],
                    "weeks": weeks,
                }
            )

        classroom_conflicts = 0
        classroom_conflict_pairs = set()  # 用于去重

        for classroom_id, schedules in classroom_schedule_with_id.items():
            # 检查每对课程是否有时间和周次重叠
            for i in range(len(schedules)):
                for j in range(i + 1, len(schedules)):
                    s1, s2 = schedules[i], schedules[j]

                    # 检查时间是否重叠
                    if s1["weekday"] != s2["weekday"]:
                        continue

                    # 检查时间段是否重叠
                    if (
                        s1["end_slot"] < s2["start_slot"]
                        or s2["end_slot"] < s1["start_slot"]
                    ):
                        continue

                    # 检查周次是否重叠
                    if not has_week_overlap(s1["weeks"], s2["weeks"]):
                        continue

                    # 确认是真冲突，检查是否已记录
                    pair_key = tuple(sorted([s1["schedule_id"], s2["schedule_id"]]))
                    if (classroom_id, pair_key) in classroom_conflict_pairs:
                        continue

                    classroom_conflict_pairs.add((classroom_id, pair_key))
                    classroom_conflicts += 1

                    day_names = [
                        "",
                        "周一",
                        "周二",
                        "周三",
                        "周四",
                        "周五",
                        "周六",
                        "周日",
                    ]
                    overlap_start = max(s1["start_slot"], s2["start_slot"])
                    overlap_end = min(s1["end_slot"], s2["end_slot"])

                    print(f"\n冲突 #{classroom_conflicts}:")
                    print(f"  教室: {s1['classroom']}")
                    print(
                        f"  时间: {day_names[s1['weekday']]} 第{overlap_start}-{overlap_end}节"
                    )
                    print(
                        f"  课程1: {s1['course']} ({s1['start_slot']}-{s1['end_slot']}节) - {s1['teacher']}"
                    )
                    print(
                        f"  课程2: {s2['course']} ({s2['start_slot']}-{s2['end_slot']}节) - {s2['teacher']}"
                    )

                    # 记录冲突数据用于导出
                    classroom_conflicts_data.append(
                        {
                            "classroom": s1["classroom"],
                            "weekday": s1["weekday"],
                            "overlap_start": overlap_start,
                            "overlap_end": overlap_end,
                            "items": [s1, s2],
                        }
                    )

        if classroom_conflicts == 0:
            print("✓ 未发现教室冲突")
        else:
            print(f"\n⚠ 共发现 {classroom_conflicts} 处教室冲突")

        print("\n" + "=" * 80)
        print("分析完成")
        print("=" * 80)

        # 导出冲突Excel
        total_conflicts = (
            conflicts_found
            + teacher_conflicts
            + classroom_conflicts
            + len(capacity_violations)
        )
        if total_conflicts > 0:
            export_conflicts_to_excel(
                class_conflicts_data,
                teacher_conflicts_data,
                classroom_conflicts_data,
                capacity_violations,
                version_id,
            )

            # 询问是否优化冲突
            print("\n" + "=" * 80)
            choice = input("是否尝试优化这些冲突? (y/n): ").strip().lower()
            if choice == "y":
                optimize_conflicts(
                    version_id,
                    results,
                    task_classes,
                    class_conflicts_data,
                    teacher_conflicts_data,
                    classroom_conflicts_data,
                    capacity_conflicts_data,
                    high_waste_data,
                    cursor,
                    conn,
                )
        else:
            # 即使没有硬约束冲突，也分析个性化要求满足情况
            print("\n" + "=" * 80)
            print("个性化要求满足情况分析")
            print("=" * 80)
            show_remaining_preference_violations(version_id, cursor)

            # 询问是否优化个性化要求
            print("\n" + "=" * 80)
            choice = input("是否尝试优化个性化要求? (y/n): ").strip().lower()
            if choice == "y":
                optimize_preferences(version_id, results, task_classes, cursor, conn)

    finally:
        cursor.close()
        conn.close()


def optimize_conflicts(
    version_id,
    results,
    task_classes,
    class_conflicts_data,
    teacher_conflicts_data,
    classroom_conflicts_data,
    capacity_conflicts_data,
    high_waste_data,
    cursor,
    conn,
):
    """优化冲突课程 - 分三阶段：1.解决容量冲突 2.解决时间冲突 3.优化利用率和个性化要求"""
    print("\n" + "=" * 80)
    print("开始优化冲突（三阶段策略）...")
    print("第一阶段：解决容量不足冲突（硬约束）")
    print("第二阶段：解决教师/班级/教室时间冲突（硬约束）")
    print("第三阶段：优化容量利用率和满足个性化要求（软约束）")
    print("=" * 80)

    # ========== 第一阶段：优化容量不足冲突（最高优先级） ==========
    capacity_adjustments = []
    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    if capacity_conflicts_data:
        print(f"\n第一阶段：解决 {len(capacity_conflicts_data)} 个容量不足冲突...")

        # 获取所有可用教室
        cursor.execute(
            "SELECT classroom_id, classroom_name, capacity, campus_id FROM classrooms ORDER BY capacity"
        )
        all_classrooms = cursor.fetchall()

        # 跟踪本批次中已分配的教室（避免重复分配）
        from collections import defaultdict

        batch_allocated = defaultdict(set)  # (weekday, slot) -> set(classroom_ids)

        for cap_conflict in capacity_conflicts_data:
            schedule_id = cap_conflict["schedule_id"]
            current_classroom_id = cap_conflict["classroom_id"]
            required_capacity = int(cap_conflict["students"])  # 转换为整数
            weekday = cap_conflict["weekday"]
            start_slot = cap_conflict["start_slot"]
            slots_count = cap_conflict["slots_count"]
            end_slot = start_slot + slots_count - 1

            print(f"\n为 {cap_conflict['course']} 寻找更大的教室...")
            print(
                f"  当前教室: {cap_conflict['classroom']} (容量{cap_conflict['capacity']})"
            )
            print(f"  需要容量: {required_capacity}")

            # 查询该时间段被占用的教室
            occupied_classrooms_query = """
            SELECT DISTINCT classroom_id 
            FROM schedules 
            WHERE version_id = %s 
                AND week_day = %s 
                AND schedule_id != %s
                AND (
                    (start_slot <= %s AND start_slot + 
                        (SELECT slots_count FROM teaching_tasks WHERE task_id = schedules.task_id) - 1 >= %s)
                    OR (start_slot >= %s AND start_slot <= %s)
                )
            """
            cursor.execute(
                occupied_classrooms_query,
                (
                    version_id,
                    weekday,
                    schedule_id,
                    start_slot,
                    start_slot,
                    start_slot,
                    end_slot,
                ),
            )
            occupied = {row["classroom_id"] for row in cursor.fetchall()}

            # 添加本批次中已分配的教室到占用列表
            for slot in range(start_slot, end_slot + 1):
                occupied.update(batch_allocated[(weekday, slot)])

            # 寻找容量足够且未被占用的教室
            suitable_classrooms = [
                room
                for room in all_classrooms
                if room["capacity"] >= required_capacity
                and room["classroom_id"] not in occupied
                and room["classroom_id"] != current_classroom_id
            ]

            if suitable_classrooms:
                # 选择容量最接近（但足够）的教室，避免浪费
                best_classroom = min(suitable_classrooms, key=lambda r: r["capacity"])
                utilization = (
                    required_capacity / best_classroom["capacity"]
                    if best_classroom["capacity"] > 0
                    else 0
                )

                capacity_adjustments.append(
                    {
                        "schedule_id": schedule_id,
                        "old_classroom_id": current_classroom_id,
                        "old_classroom": cap_conflict["classroom"],
                        "old_capacity": cap_conflict["capacity"],
                        "new_classroom_id": best_classroom["classroom_id"],
                        "new_classroom": best_classroom["classroom_name"],
                        "new_capacity": best_classroom["capacity"],
                        "students": required_capacity,
                        "utilization": utilization,
                        "course": cap_conflict["course"],
                        "time": f"{day_names[weekday]} 第{start_slot}-{end_slot}节",
                    }
                )

                # 标记本批次中该教室在这些时间槽已被占用
                for slot in range(start_slot, end_slot + 1):
                    batch_allocated[(weekday, slot)].add(best_classroom["classroom_id"])

                print(
                    f"  ✓ 找到合适教室: {best_classroom['classroom_name']} (容量{best_classroom['capacity']}, 利用率{utilization:.1%})"
                )
            else:
                print(f"  ✗ 未找到合适的教室（该时间段容量足够的教室都已被占用）")

        # 应用容量调整
        if capacity_adjustments:
            print(f"\n找到 {len(capacity_adjustments)} 个容量优化方案:")
            for i, adj in enumerate(capacity_adjustments, 1):
                print(
                    f"  [{i}] {adj['course']}: {adj['time']}\n"
                    f"      {adj['old_classroom']} (容量{adj['old_capacity']}) → "
                    f"{adj['new_classroom']} (容量{adj['new_capacity']}, 利用率{adj['utilization']:.1%})"
                )

            confirm = input("\n确认应用这些容量优化? (y/n): ").strip().lower()
            if confirm == "y":
                for adj in capacity_adjustments:
                    update_query = (
                        "UPDATE schedules SET classroom_id = %s WHERE schedule_id = %s"
                    )
                    cursor.execute(
                        update_query, (adj["new_classroom_id"], adj["schedule_id"])
                    )

                conn.commit()
                print(f"✓ 已应用 {len(capacity_adjustments)} 个容量优化")
            else:
                print("已取消容量优化")
        else:
            print("\n⚠ 未找到可行的容量优化方案")
    else:
        print("\n第一阶段：✓ 无容量不足冲突")

    # ========== 第二阶段：优化时间冲突 ==========
    # 收集所有有时间冲突的 schedule_id
    conflict_schedule_ids = set()

    # 从班级冲突中提取（items中已包含schedule_id）
    for conflict in class_conflicts_data:
        for item in conflict["items"]:
            if "schedule_id" in item:
                conflict_schedule_ids.add(item["schedule_id"])

    # 从教师冲突中提取
    for conflict in teacher_conflicts_data:
        for item in conflict["items"]:
            if "schedule_id" in item:
                conflict_schedule_ids.add(item["schedule_id"])

    # 从教室冲突中提取
    for conflict in classroom_conflicts_data:
        for item in conflict["items"]:
            if "schedule_id" in item:
                conflict_schedule_ids.add(item["schedule_id"])

    print(f"\n第二阶段：识别到 {len(conflict_schedule_ids)} 个有时间冲突的课程安排")

    if len(conflict_schedule_ids) == 0:
        print("✓ 没有硬约束冲突")
        # 即使没有硬约束冲突，也可以尝试优化个性化要求
        optimize_preferences(version_id, results, task_classes, cursor, conn)
        return

    # 获取已占用的时间槽(不包括待调整的课程)
    occupied_times = defaultdict(
        lambda: defaultdict(set)
    )  # [entity_type][entity_id] = set of (weekday, slot)

    for result in results:
        if result["schedule_id"] in conflict_schedule_ids:
            continue  # 跳过冲突课程

        task_id = result["task_id"]
        start_slot = result["start_slot"]
        end_slot = start_slot + result["slots_count"] - 1
        weekday = result["week_day"]
        classroom_id = result["classroom_id"]

        # 记录教室占用
        for slot in range(start_slot, end_slot + 1):
            occupied_times["classroom"][classroom_id].add((weekday, slot))

        # 记录教师占用
        teacher_ids_str = result.get("teacher_ids", "")
        if teacher_ids_str:
            for teacher_id in teacher_ids_str.split(", "):
                for slot in range(start_slot, end_slot + 1):
                    occupied_times["teacher"][teacher_id].add((weekday, slot))

        # 记录班级占用
        for cls in task_classes[task_id]:
            class_id = cls["class_id"]
            for slot in range(start_slot, end_slot + 1):
                occupied_times["class"][class_id].add((weekday, slot))

    # ========== 第一阶段：解决硬约束冲突 ==========
    print("\n正在为有冲突的课程寻找无冲突时间槽...")
    adjustments = []
    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    for schedule_id in conflict_schedule_ids:
        # 找到对应的排课记录
        schedule_record = None
        for result in results:
            if result["schedule_id"] == schedule_id:
                schedule_record = result
                break

        if not schedule_record:
            continue

        task_id = schedule_record["task_id"]
        slots_count = schedule_record["slots_count"]
        classroom_id = schedule_record["classroom_id"]
        old_weekday = schedule_record["week_day"]
        old_start_slot = schedule_record["start_slot"]

        # 获取教师和班级
        teacher_ids = schedule_record.get("teacher_ids", "").split(", ")
        teacher_ids = [tid for tid in teacher_ids if tid]  # 过滤空字符串
        classes = [cls["class_id"] for cls in task_classes[task_id]]

        print(f"\n正在为 {schedule_record['course_name']} 寻找新时间...")
        print(
            f"  当前: {day_names[old_weekday]} 第{old_start_slot}-{old_start_slot+slots_count-1}节"
        )
        print(
            f"  教师: {len(teacher_ids)}个, 班级: {len(classes)}个, 教室: {classroom_id}"
        )

        # 尝试找到新的时间槽
        found_slot = False
        time_slots = []

        # 生成可能的时间槽 (周一到周五, 1-13节)
        for weekday in range(1, 6):  # 周一到周五
            for start_slot in range(1, 14 - slots_count + 1):
                # 跳过周四下午
                if weekday == 4 and start_slot >= 6:
                    continue

                # 检查是否有冲突
                has_conflict = False

                # 检查教室冲突
                for slot in range(start_slot, start_slot + slots_count):
                    if (weekday, slot) in occupied_times["classroom"][classroom_id]:
                        has_conflict = True
                        break

                if has_conflict:
                    continue

                # 检查教师冲突
                for teacher_id in teacher_ids:
                    for slot in range(start_slot, start_slot + slots_count):
                        if (weekday, slot) in occupied_times["teacher"][teacher_id]:
                            has_conflict = True
                            break
                    if has_conflict:
                        break

                if has_conflict:
                    continue

                # 检查班级冲突
                for class_id in classes:
                    for slot in range(start_slot, start_slot + slots_count):
                        if (weekday, slot) in occupied_times["class"][class_id]:
                            has_conflict = True
                            break
                    if has_conflict:
                        break

                if not has_conflict:
                    time_slots.append((weekday, start_slot))

        print(f"  找到 {len(time_slots)} 个候选时间槽")

        if time_slots:
            # 优先选择不同于原时间的槽位
            new_weekday, new_start_slot = None, None
            for weekday, start_slot in time_slots:
                if weekday != old_weekday or start_slot != old_start_slot:
                    new_weekday, new_start_slot = weekday, start_slot
                    break

            # 如果所有时间槽都是原时间，说明该课程可以留在原位
            # 这种情况下，跳过该课程，让其他冲突课程调整
            if new_weekday is None:
                print(f"  → 保留在原位，等待其他冲突课程调整")
                continue  # 更新占用记录
            for slot in range(new_start_slot, new_start_slot + slots_count):
                occupied_times["classroom"][classroom_id].add((new_weekday, slot))
                for teacher_id in teacher_ids:
                    occupied_times["teacher"][teacher_id].add((new_weekday, slot))
                for class_id in classes:
                    occupied_times["class"][class_id].add((new_weekday, slot))

            adjustments.append(
                {
                    "schedule_id": schedule_id,
                    "course_name": schedule_record["course_name"],
                    "old_time": (old_weekday, old_start_slot),
                    "new_time": (new_weekday, new_start_slot),
                    "new_end_slot": new_start_slot + slots_count - 1,
                }
            )
        else:
            print(f"⚠ 无法为课程 {schedule_record['course_name']} 找到合适的时间槽")
            print(f"   原时间: {day_names[old_weekday]} 第{old_start_slot}节")
            print(
                f"   课程信息: {slots_count}节课, 教室{classroom_id}, {len(teacher_ids)}个教师, {len(classes)}个班级"
            )

            # 诊断：检查是否有可用时间（不考虑教室和教师限制）
            available_for_class = []
            day_names_local = ["", "周一", "周二", "周三", "周四", "周五"]
            for weekday in range(1, 6):
                for start_slot in range(1, 14 - slots_count + 1):
                    if weekday == 4 and start_slot >= 6:
                        continue

                    # 只检查班级冲突
                    has_class_conflict = False
                    for class_id in classes:
                        for slot in range(start_slot, start_slot + slots_count):
                            if (weekday, slot) in occupied_times["class"][class_id]:
                                has_class_conflict = True
                                break
                        if has_class_conflict:
                            break

                    if not has_class_conflict and (
                        weekday != old_weekday or start_slot != old_start_slot
                    ):
                        available_for_class.append(
                            f"{day_names_local[weekday]}{start_slot}-{start_slot+slots_count-1}"
                        )

            if available_for_class:
                print(
                    f"   班级空闲时间: {', '.join(available_for_class[:5])}"
                    + (
                        f" 等{len(available_for_class)}个"
                        if len(available_for_class) > 5
                        else ""
                    )
                )
                print(f"   → 可能是教师或教室时间冲突导致无法调整")

    if adjustments:
        print(f"\n找到 {len(adjustments)} 个调整方案:")
        day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        for adj in adjustments:
            old_day, old_slot = adj["old_time"]
            new_day, new_slot = adj["new_time"]
            print(
                f"  {adj['course_name']}: {day_names[old_day]} 第{old_slot}节 -> {day_names[new_day]} 第{new_slot}节"
            )

        confirm = input("\n确认应用这些调整? (y/n): ").strip().lower()
        if confirm == "y":
            # 应用调整到数据库
            for adj in adjustments:
                update_query = """
                UPDATE schedules 
                SET week_day = %s, start_slot = %s, end_slot = %s
                WHERE schedule_id = %s
                """
                cursor.execute(
                    update_query,
                    (
                        adj["new_time"][0],
                        adj["new_time"][1],
                        adj["new_end_slot"],
                        adj["schedule_id"],
                    ),
                )

            conn.commit()
            print(
                f"\n✓ 第二阶段完成：已成功调整 {len(adjustments)} 个课程安排，解决时间冲突"
            )

            # 第三阶段：优化容量利用率和个性化要求
            print("\n" + "=" * 80)
            print("第三阶段：优化容量利用率和个性化要求")
            print("=" * 80)
            optimize_utilization_and_preferences(
                version_id, results, task_classes, high_waste_data, cursor, conn
            )

            # 优化完成后，重新生成冲突报告
            print("\n" + "=" * 80)
            print("重新生成优化后的冲突报告...")
            print("=" * 80)
            regenerate_conflict_report_after_optimization(version_id, cursor)
        else:
            print("\n已取消调整")
    else:
        print("\n⚠ 未找到可行的调整方案解决所有硬约束冲突")
        print("建议：")
        print("  1. 检查教室资源是否充足")
        print("  2. 检查教师时间是否过于紧张")
        print("  3. 考虑调整课程节数或增加教室/教师资源")


def optimize_utilization_and_preferences(
    version_id, results, task_classes, high_waste_data, cursor, conn
):
    """第三阶段：优化容量利用率，然后优化个性化要求"""

    # 先优化利用率
    if high_waste_data:
        print(f"\n发现 {len(high_waste_data)} 个容量浪费严重的课程，尝试优化...")

        # 获取所有教室
        cursor.execute(
            "SELECT classroom_id, classroom_name, capacity, campus_id FROM classrooms ORDER BY capacity"
        )
        all_classrooms = cursor.fetchall()

        utilization_adjustments = []
        day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

        # 跟踪本批次中已分配的教室（避免重复分配）
        # allocated_classrooms[(weekday, slot)] = set(classroom_ids)
        # 使用单个时间槽而不是时间段，以便正确检测重叠
        from collections import defaultdict

        allocated_classrooms = defaultdict(set)

        for waste in high_waste_data[:20]:  # 限制优化前20个最严重的
            schedule_id = waste["schedule_id"]
            current_classroom_id = waste["classroom_id"]
            student_count = int(waste["students"])  # 转换为整数
            weekday = waste["weekday"]
            start_slot = waste["start_slot"]
            slots_count = waste["slots_count"]
            end_slot = start_slot + slots_count - 1

            # 查询该时间段被占用的教室（包括数据库中的和本批次中已分配的）
            occupied_classrooms_query = """
            SELECT DISTINCT classroom_id 
            FROM schedules 
            WHERE version_id = %s 
                AND week_day = %s 
                AND schedule_id != %s
                AND (
                    (start_slot <= %s AND start_slot + 
                        (SELECT slots_count FROM teaching_tasks WHERE task_id = schedules.task_id) - 1 >= %s)
                    OR (start_slot >= %s AND start_slot <= %s)
                )
            """
            cursor.execute(
                occupied_classrooms_query,
                (
                    version_id,
                    weekday,
                    schedule_id,
                    start_slot,
                    start_slot,
                    start_slot,
                    end_slot,
                ),
            )
            occupied = {row["classroom_id"] for row in cursor.fetchall()}

            # 合并本批次中已分配的教室（检查所有重叠的时间槽）
            for slot in range(start_slot, end_slot + 1):
                occupied.update(allocated_classrooms[(weekday, slot)])

            # 寻找容量更合适的教室（利用率在80%-95%之间）
            target_min_capacity = int(student_count / 0.95)  # 最小容量（达到95%利用率）
            target_max_capacity = int(student_count / 0.80)  # 最大容量（达到80%利用率）

            suitable_classrooms = [
                room
                for room in all_classrooms
                if target_min_capacity <= room["capacity"] <= target_max_capacity
                and room["classroom_id"] not in occupied
                and room["classroom_id"] != current_classroom_id
            ]

            if suitable_classrooms:
                # 选择利用率最接近87.5%的教室
                best_classroom = min(
                    suitable_classrooms,
                    key=lambda r: (
                        abs(student_count / r["capacity"] - 0.875)
                        if r["capacity"] > 0
                        else 999
                    ),
                )
                new_utilization = (
                    student_count / best_classroom["capacity"]
                    if best_classroom["capacity"] > 0
                    else 0
                )

                # 只有当新利用率明显更好时才调整
                if new_utilization > waste["waste_ratio"] + 0.1:  # 至少提升10%
                    utilization_adjustments.append(
                        {
                            "schedule_id": schedule_id,
                            "course": waste["course"],
                            "old_classroom": waste["classroom"],
                            "old_capacity": waste["capacity"],
                            "old_utilization": 1 - waste["waste_ratio"],
                            "new_classroom_id": best_classroom["classroom_id"],
                            "new_classroom": best_classroom["classroom_name"],
                            "new_capacity": best_classroom["capacity"],
                            "new_utilization": new_utilization,
                            "students": student_count,
                            "time": f"{day_names[weekday]} 第{start_slot}-{end_slot}节",
                        }
                    )
                    # 标记该教室在此时间段的所有时间槽已被本批次分配
                    for slot in range(start_slot, end_slot + 1):
                        allocated_classrooms[(weekday, slot)].add(
                            best_classroom["classroom_id"]
                        )

        if utilization_adjustments:
            print(f"\n找到 {len(utilization_adjustments)} 个利用率优化方案:")
            for i, adj in enumerate(utilization_adjustments, 1):
                print(
                    f"  [{i}] {adj['course']}: {adj['time']}\n"
                    f"      {adj['old_classroom']} (容量{adj['old_capacity']}, 利用率{adj['old_utilization']:.1%}) → "
                    f"{adj['new_classroom']} (容量{adj['new_capacity']}, 利用率{adj['new_utilization']:.1%})"
                )

            confirm = input("\n确认应用这些利用率优化? (y/n): ").strip().lower()
            if confirm == "y":
                for adj in utilization_adjustments:
                    update_query = (
                        "UPDATE schedules SET classroom_id = %s WHERE schedule_id = %s"
                    )
                    cursor.execute(
                        update_query, (adj["new_classroom_id"], adj["schedule_id"])
                    )

                conn.commit()
                print(f"✓ 已应用 {len(utilization_adjustments)} 个利用率优化")
            else:
                print("已取消利用率优化")
        else:
            print("\n未找到可以进一步优化利用率的方案")
    else:
        print("\n✓ 容量利用率已经很好，无需优化")

    # 然后优化个性化要求
    print("\n" + "-" * 80)
    print("优化个性化要求（教师偏好时间）")
    print("-" * 80)
    optimize_preferences(version_id, results, task_classes, cursor, conn)

    # 最后优化任务关系约束
    print("\n" + "-" * 80)
    print("优化任务关系约束")
    print("-" * 80)
    optimize_task_relations(version_id, results, task_classes, cursor, conn)


def optimize_preferences(version_id, results, task_classes, cursor, conn):
    """第二阶段：在不违反硬约束的前提下，优化个性化要求（教师偏好时间）"""
    from collections import defaultdict

    # 加载offering_weeks表数据
    cursor.execute("SELECT offering_id, week_number FROM offering_weeks")
    offering_weeks = {}
    for row in cursor.fetchall():
        offering_id = row["offering_id"]
        week_number = row["week_number"]
        if offering_id not in offering_weeks:
            offering_weeks[offering_id] = set()
        offering_weeks[offering_id].add(week_number)

    # 定义获取周次的函数
    def get_weeks(result):
        """根据offering_id获取周次集合"""
        offering_id = result["offering_id"]
        if offering_id in offering_weeks:
            return offering_weeks[offering_id]

        # 如果没有自定义周次,根据week_pattern生成
        start_week = result.get("start_week") or 1
        end_week = result.get("end_week") or 17
        week_pattern = result.get("week_pattern") or "CONTINUOUS"

        if week_pattern == "CONTINUOUS":
            return set(range(start_week, end_week + 1))
        elif week_pattern == "SINGLE":
            return set(w for w in range(start_week, end_week + 1) if w % 2 == 1)
        elif week_pattern == "DOUBLE":
            return set(w for w in range(start_week, end_week + 1) if w % 2 == 0)
        else:
            return set(range(start_week, end_week + 1))

    # 获取教师偏好设置
    pref_query = """
    SELECT 
        tp.preference_id,
        tp.offering_id,
        tp.teacher_id,
        tp.preference_type,
        tp.weekday,
        tp.start_slot,
        tp.end_slot,
        tp.penalty_score,
        t.teacher_name
    FROM teacher_preferences tp
    JOIN teachers t ON tp.teacher_id = t.teacher_id
    WHERE tp.offering_id IN (
        SELECT DISTINCT tt.offering_id 
        FROM teaching_tasks tt
        JOIN schedules s ON tt.task_id = s.task_id
        WHERE s.version_id = %s
    )
    ORDER BY tp.penalty_score DESC, tp.preference_type
    """

    cursor.execute(pref_query, (version_id,))
    preferences = cursor.fetchall()

    if not preferences:
        print("\n✓ 没有设置教师个性化要求")
        return

    print(f"\n找到 {len(preferences)} 条教师个性化要求")

    # 分类统计
    preferred_prefs = [p for p in preferences if p["preference_type"] == "PREFERRED"]
    avoided_prefs = [p for p in preferences if p["preference_type"] == "AVOIDED"]

    print(f"  - 偏好时间（PREFERRED）: {len(preferred_prefs)} 条")
    print(f"  - 避免时间（AVOIDED）: {len(avoided_prefs)} 条")

    # 构建当前占用情况（用于检查硬约束）- 包含周次信息
    # occupied_times[type][id][(weekday, slot)] = [weeks1, weeks2, ...]
    occupied_times = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    schedule_map = {}  # schedule_id -> result

    for result in results:
        schedule_id = result["schedule_id"]
        task_id = result["task_id"]
        start_slot = result["start_slot"]
        end_slot = start_slot + result["slots_count"] - 1
        weekday = result["week_day"]
        classroom_id = result["classroom_id"]

        # 获取周次信息
        weeks = get_weeks(result)

        schedule_map[schedule_id] = result

        # 记录占用（包含周次）
        for slot in range(start_slot, end_slot + 1):
            time_key = (weekday, slot)
            occupied_times["classroom"][classroom_id][time_key].append(weeks)

            teacher_ids_str = result.get("teacher_ids", "")
            if teacher_ids_str:
                for teacher_id in teacher_ids_str.split(", "):
                    occupied_times["teacher"][teacher_id][time_key].append(weeks)

            for cls in task_classes[task_id]:
                occupied_times["class"][cls["class_id"]][time_key].append(weeks)

    # 分析当前个性化要求满足情况
    print("\n分析当前个性化要求满足情况...")

    violated_preferred = []  # 未满足的偏好时间
    violated_avoided = []  # 违反的避免时间

    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    for pref in preferences:
        teacher_id = pref["teacher_id"]
        teacher_name = pref["teacher_name"]
        pref_type = pref["preference_type"]
        weekday = pref["weekday"]
        start_slot = pref["start_slot"]
        end_slot = pref["end_slot"]

        if not weekday or not start_slot or not end_slot:
            continue

        # 查找该教师在该偏好时间段的课程
        teacher_schedule = occupied_times["teacher"][teacher_id]

        has_class_in_range = False
        for wd, slot in teacher_schedule:
            if wd == weekday and start_slot <= slot <= end_slot:
                has_class_in_range = True
                break

        if pref_type == "PREFERRED" and not has_class_in_range:
            violated_preferred.append(pref)
        elif pref_type == "AVOIDED" and has_class_in_range:
            violated_avoided.append(pref)

    print(f"\n个性化要求满足情况：")
    print(
        f"  ✓ 偏好时间已满足: {len(preferred_prefs) - len(violated_preferred)}/{len(preferred_prefs)}"
    )
    print(
        f"  ✓ 避免时间已遵守: {len(avoided_prefs) - len(violated_avoided)}/{len(avoided_prefs)}"
    )

    if violated_preferred:
        print(f"\n  ⚠ 未满足的偏好时间: {len(violated_preferred)} 条")
        for i, pref in enumerate(violated_preferred[:5], 1):
            print(
                f"    [{i}] {pref['teacher_name']}: {day_names[pref['weekday']]} {pref['start_slot']}-{pref['end_slot']}节"
            )
        if len(violated_preferred) > 5:
            print(f"    ... 还有 {len(violated_preferred) - 5} 条")

    if violated_avoided:
        print(f"\n  ⚠ 违反的避免时间: {len(violated_avoided)} 条")
        for i, pref in enumerate(violated_avoided[:5], 1):
            print(
                f"    [{i}] {pref['teacher_name']}: {day_names[pref['weekday']]} {pref['start_slot']}-{pref['end_slot']}节"
            )
        if len(violated_avoided) > 5:
            print(f"    ... 还有 {len(violated_avoided) - 5} 条")

    # 尝试优化个性化要求
    total_violations = len(violated_preferred) + len(violated_avoided)
    if total_violations == 0:
        print("\n✓ 所有个性化要求都已满足，无需优化")
        return

    print(f"\n尝试优化 {total_violations} 条未满足的个性化要求...")
    print("注意：只会调整不违反硬约束（教师/班级/教室冲突）的课程")

    preference_adjustments = []

    # 优先处理违反避免时间的情况（penalty更高）
    for pref in sorted(
        violated_avoided, key=lambda x: x["penalty_score"], reverse=True
    ):
        # 找到该教师在该时间段的课程
        teacher_id = pref["teacher_id"]
        weekday = pref["weekday"]
        start_slot = pref["start_slot"]
        end_slot = pref["end_slot"]

        # 查找需要调整的schedule
        for schedule_id, result in schedule_map.items():
            teacher_ids_str = result.get("teacher_ids", "")
            if teacher_id not in teacher_ids_str.split(", "):
                continue

            if result["week_day"] != weekday:
                continue

            # 检查时间是否重叠
            result_start = result["start_slot"]
            result_end = result_start + result["slots_count"] - 1

            if not (result_end < start_slot or result_start > end_slot):
                # 找到冲突课程，尝试调整
                new_time = find_alternative_time_for_preference(
                    result,
                    task_classes,
                    occupied_times,
                    schedule_id,
                    avoid_time=(weekday, start_slot, end_slot),
                    get_weeks_func=get_weeks,
                )

                if new_time:
                    preference_adjustments.append(
                        {
                            "schedule_id": schedule_id,
                            "course_name": result["course_name"],
                            "teacher_name": pref["teacher_name"],
                            "old_time": (result["week_day"], result["start_slot"]),
                            "new_time": new_time,
                            "reason": f"避免时间 {day_names[weekday]} {start_slot}-{end_slot}节",
                            "result": result,
                        }
                    )

    # 尝试满足偏好时间
    for pref in sorted(
        violated_preferred, key=lambda x: x["penalty_score"], reverse=True
    )[
        :10
    ]:  # 限制处理数量
        teacher_id = pref["teacher_id"]
        weekday = pref["weekday"]
        start_slot = pref["start_slot"]
        end_slot = pref["end_slot"]

        # 找到该教师的课程，尝试移到偏好时间
        for schedule_id, result in schedule_map.items():
            teacher_ids_str = result.get("teacher_ids", "")
            if teacher_id not in teacher_ids_str.split(", "):
                continue

            # 如果已经在调整列表中，跳过
            if any(adj["schedule_id"] == schedule_id for adj in preference_adjustments):
                continue

            # 尝试移到偏好时间段
            new_time = find_alternative_time_for_preference(
                result,
                task_classes,
                occupied_times,
                schedule_id,
                prefer_time=(weekday, start_slot, end_slot),
                get_weeks_func=get_weeks,
            )

            if (
                new_time
                and new_time[0] == weekday
                and start_slot <= new_time[1] <= end_slot
            ):
                preference_adjustments.append(
                    {
                        "schedule_id": schedule_id,
                        "course_name": result["course_name"],
                        "teacher_name": pref["teacher_name"],
                        "old_time": (result["week_day"], result["start_slot"]),
                        "new_time": new_time,
                        "reason": f"满足偏好时间 {day_names[weekday]} {start_slot}-{end_slot}节",
                        "result": result,
                    }
                )
                break  # 每个偏好只调整一门课

    if preference_adjustments:
        print(f"\n找到 {len(preference_adjustments)} 个可优化的课程安排：")
        for i, adj in enumerate(preference_adjustments, 1):
            old_day, old_slot = adj["old_time"]
            new_day, new_slot = adj["new_time"]
            print(f"  [{i}] {adj['course_name']} ({adj['teacher_name']})")
            print(
                f"      {day_names[old_day]} 第{old_slot}节 -> {day_names[new_day]} 第{new_slot}节"
            )
            print(f"      原因: {adj['reason']}")

        confirm = input("\n确认应用这些优化? (y/n): ").strip().lower()
        if confirm == "y":
            for adj in preference_adjustments:
                result = adj["result"]
                new_day, new_slot = adj["new_time"]
                new_end_slot = new_slot + result["slots_count"] - 1

                update_query = """
                UPDATE schedules 
                SET week_day = %s, start_slot = %s, end_slot = %s
                WHERE schedule_id = %s
                """
                cursor.execute(
                    update_query, (new_day, new_slot, new_end_slot, adj["schedule_id"])
                )

            conn.commit()
            print(
                f"\n✓ 第二阶段完成：已优化 {len(preference_adjustments)} 个课程以满足个性化要求"
            )

            # 重新分析优化后的个性化要求满足情况
            print("\n" + "=" * 80)
            print("重新分析优化后的个性化要求满足情况...")
            print("=" * 80)
            show_remaining_preference_violations(version_id, cursor)

            print("\n建议重新运行冲突分析验证结果")
        else:
            print("\n已取消优化")
    else:
        print("\n未找到可以优化的方案（所有调整都会违反硬约束）")

        # 即使无法优化，也显示当前未满足的个性化要求
        if total_violations > 0:
            print("\n" + "=" * 80)
            print("当前仍未满足的个性化要求详情：")
            print("=" * 80)
            show_remaining_preference_violations(version_id, cursor)


def optimize_task_relations(version_id, results, task_classes, cursor, conn):
    """优化任务关系约束违反情况"""
    from collections import defaultdict

    print("\n" + "-" * 80)
    print("优化任务关系约束（课程多次上课的时间关系）")
    print("-" * 80)

    # 获取任务关系约束
    query = """
    SELECT 
        trc.*,
        tta.task_id AS task_id_a,
        ttb.task_id AS task_id_b,
        c.course_name
    FROM task_relation_constraints trc
    JOIN course_offerings co ON trc.offering_id = co.offering_id
    JOIN courses c ON co.course_id = c.course_id
    JOIN teaching_tasks tta ON co.offering_id = tta.offering_id 
        AND tta.task_sequence = trc.task_sequence_a
    JOIN teaching_tasks ttb ON co.offering_id = ttb.offering_id 
        AND ttb.task_sequence = trc.task_sequence_b
    WHERE co.offering_id IN (
        SELECT DISTINCT tt.offering_id 
        FROM teaching_tasks tt
        JOIN schedules s ON tt.task_id = s.task_id
        WHERE s.version_id = %s
    )
    """
    cursor.execute(query, (version_id,))
    constraints = cursor.fetchall()

    if not constraints:
        print("\n✓ 未设置任务关系约束")
        return

    # 加载offering_weeks数据
    cursor.execute("SELECT offering_id, week_number FROM offering_weeks")
    offering_weeks = {}
    for row in cursor.fetchall():
        offering_id = row["offering_id"]
        week_number = row["week_number"]
        if offering_id not in offering_weeks:
            offering_weeks[offering_id] = set()
        offering_weeks[offering_id].add(week_number)

    def get_weeks(result):
        """根据offering_id获取周次集合"""
        offering_id = result["offering_id"]
        if offering_id in offering_weeks:
            return offering_weeks[offering_id]
        start_week = result.get("start_week") or 1
        end_week = result.get("end_week") or 17
        week_pattern = result.get("week_pattern") or "CONTINUOUS"
        if week_pattern == "CONTINUOUS":
            return set(range(start_week, end_week + 1))
        elif week_pattern == "SINGLE":
            return set(w for w in range(start_week, end_week + 1) if w % 2 == 1)
        elif week_pattern == "DOUBLE":
            return set(w for w in range(start_week, end_week + 1) if w % 2 == 0)
        else:
            return set(range(start_week, end_week + 1))

    # 构建当前占用情况
    occupied_times = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    schedule_map = {}

    for result in results:
        schedule_id = result["schedule_id"]
        schedule_map[schedule_id] = result
        task_id = result["task_id"]
        start_slot = result["start_slot"]
        end_slot = start_slot + result["slots_count"] - 1
        weekday = result["week_day"]
        classroom_id = result["classroom_id"]
        weeks = get_weeks(result)

        for slot in range(start_slot, end_slot + 1):
            time_key = (weekday, slot)
            occupied_times["classroom"][classroom_id][time_key].append(weeks)
            teacher_ids_str = result.get("teacher_ids", "")
            if teacher_ids_str:
                for teacher_id in teacher_ids_str.split(", "):
                    if teacher_id:
                        occupied_times["teacher"][teacher_id][time_key].append(weeks)
            for cls in task_classes[task_id]:
                occupied_times["class"][cls["class_id"]][time_key].append(weeks)

    # 分析违反的约束
    violations = []
    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    for const in constraints:
        cursor.execute(
            "SELECT * FROM schedules WHERE version_id = %s AND task_id = %s",
            (version_id, const["task_id_a"]),
        )
        sch_a = cursor.fetchone()
        cursor.execute(
            "SELECT * FROM schedules WHERE version_id = %s AND task_id = %s",
            (version_id, const["task_id_b"]),
        )
        sch_b = cursor.fetchone()

        if not sch_a or not sch_b:
            continue

        # 找到完整的result对象
        result_a = schedule_map.get(sch_a["schedule_id"])
        result_b = schedule_map.get(sch_b["schedule_id"])

        if not result_a or not result_b:
            continue

        day_diff = abs(sch_a["week_day"] - sch_b["week_day"])
        constraint_type = const["constraint_type"]
        violated = False

        if constraint_type == "REQUIRE_SAME_DAY":
            if sch_a["week_day"] != sch_b["week_day"]:
                violated = True
        elif constraint_type == "AVOID_CONSECUTIVE_DAYS":
            if day_diff == 1:
                violated = True
        elif constraint_type == "MIN_DAYS_APART":
            min_gap = const["constraint_value"] or 1
            if day_diff < min_gap:
                violated = True

        if violated:
            violations.append(
                {
                    "constraint": const,
                    "result_a": result_a,
                    "result_b": result_b,
                    "sch_a": sch_a,
                    "sch_b": sch_b,
                }
            )

    if not violations:
        print("\n✓ 所有任务关系约束均已满足")
        return

    print(f"\n找到 {len(violations)} 处任务关系约束违反")

    # 尝试优化
    adjustments = []

    for v in violations:
        const = v["constraint"]
        result_a = v["result_a"]
        result_b = v["result_b"]
        sch_a = v["sch_a"]
        sch_b = v["sch_b"]

        print(f"\n优化: {const['course_name']} - {const['constraint_type']}")
        print(
            f"  当前: 任务A={day_names[sch_a['week_day']]} 第{sch_a['start_slot']}节, "
            f"任务B={day_names[sch_b['week_day']]} 第{sch_b['start_slot']}节"
        )

        # 根据约束类型寻找合适的时间
        target_weekday = None
        if const["constraint_type"] == "REQUIRE_SAME_DAY":
            # 尝试将任务B调整到与任务A同一天
            target_weekday = sch_a["week_day"]
            print(f"  目标: 将任务B调整到 {day_names[target_weekday]}")

        elif const["constraint_type"] == "AVOID_CONSECUTIVE_DAYS":
            # 尝试将任务B调整到非连续天
            current_day = sch_b["week_day"]
            adjacent_days = {sch_a["week_day"] - 1, sch_a["week_day"] + 1}
            for day in range(1, 6):  # 周一到周五
                if day not in adjacent_days:
                    target_weekday = day
                    break
            if target_weekday:
                print(f"  目标: 将任务B调整到非连续天 {day_names[target_weekday]}")

        elif const["constraint_type"] == "MIN_DAYS_APART":
            # 尝试增加天数间隔
            min_gap = const["constraint_value"] or 1
            target_weekday = sch_a["week_day"] + min_gap
            if target_weekday > 5:
                target_weekday = sch_a["week_day"] - min_gap
            if 1 <= target_weekday <= 5:
                print(
                    f"  目标: 将任务B调整到间隔≥{min_gap}天的 {day_names[target_weekday]}"
                )

        if not target_weekday or target_weekday < 1 or target_weekday > 5:
            print("  ✗ 无法找到合适的目标时间")
            continue

        # 查找任务B在目标星期的可用时间槽
        found_slot = False
        task_b_id = sch_b["task_id"]
        task_b_obj = next(
            (t for r in results if r["task_id"] == task_b_id for t in [r]), None
        )
        if not task_b_obj:
            continue

        slots_count = task_b_obj["slots_count"]
        classroom_id = sch_b["classroom_id"]
        teacher_ids = result_b.get("teacher_ids", "").split(", ")
        teacher_ids = [tid for tid in teacher_ids if tid]
        class_ids = [cls["class_id"] for cls in task_classes[task_b_id]]
        weeks_b = get_weeks(result_b)

        for start_slot in range(1, 14 - slots_count + 1):
            if check_time_available(
                target_weekday,
                start_slot,
                slots_count,
                classroom_id,
                teacher_ids,
                class_ids,
                occupied_times,
                weeks_b,
            ):
                adjustments.append(
                    {
                        "schedule_id": sch_b["schedule_id"],
                        "task_id": task_b_id,
                        "old_weekday": sch_b["week_day"],
                        "old_start": sch_b["start_slot"],
                        "new_weekday": target_weekday,
                        "new_start": start_slot,
                        "course": const["course_name"],
                        "reason": f"满足任务关系约束: {const['constraint_type']}",
                    }
                )
                print(
                    f"  ✓ 找到可用时间: {day_names[target_weekday]} 第{start_slot}-{start_slot+slots_count-1}节"
                )
                found_slot = True
                break

        if not found_slot:
            print("  ✗ 未找到可用时间槽（所有时间都有冲突）")

    if adjustments:
        print(f"\n找到 {len(adjustments)} 个可优化的任务关系约束")
        print("\n调整详情:")
        for i, adj in enumerate(adjustments, 1):
            print(f"  [{i}] {adj['course']}")
            print(f"      {adj['reason']}")
            print(
                f"      {day_names[adj['old_weekday']]} 第{adj['old_start']}节 → "
                f"{day_names[adj['new_weekday']]} 第{adj['new_start']}节"
            )

        choice = input("\n是否应用这些调整? (y/n): ").strip().lower()
        if choice == "y":
            for adj in adjustments:
                cursor.execute(
                    """
                    UPDATE schedules 
                    SET week_day = %s, start_slot = %s 
                    WHERE schedule_id = %s
                    """,
                    (adj["new_weekday"], adj["new_start"], adj["schedule_id"]),
                )
            conn.commit()
            print(f"\n✅ 已成功优化 {len(adjustments)} 处任务关系约束违反")
            print("建议重新运行冲突分析验证结果")
        else:
            print("\n已取消优化")
    else:
        print("\n未找到可以优化的方案（所有调整都会违反硬约束）")


def find_alternative_time_for_preference(
    result,
    task_classes,
    occupied_times,
    current_schedule_id,
    avoid_time=None,
    prefer_time=None,
    get_weeks_func=None,
):
    """为课程寻找替代时间，考虑个性化要求

    Args:
        result: 当前排课记录
        task_classes: 任务-班级映射
        occupied_times: 已占用时间
        current_schedule_id: 当前schedule_id（需要从占用时间中排除）
        avoid_time: 要避免的时间段 (weekday, start_slot, end_slot)
        prefer_time: 偏好的时间段 (weekday, start_slot, end_slot)
        get_weeks_func: 获取周次的函数
    """
    task_id = result["task_id"]
    slots_count = result["slots_count"]
    classroom_id = result["classroom_id"]
    teacher_ids = result.get("teacher_ids", "").split(", ")
    teacher_ids = [tid for tid in teacher_ids if tid]
    classes = [cls["class_id"] for cls in task_classes[task_id]]

    # 获取周次信息
    weeks = get_weeks_func(result) if get_weeks_func else None

    # 临时移除当前课程的占用
    old_weekday = result["week_day"]
    old_start = result["start_slot"]
    old_end = old_start + slots_count - 1

    # 由于occupied_times现在包含周次信息，需要小心处理
    # 暂时跳过移除和恢复的步骤，直接搜索可用时间
    # TODO: 需要更仔细的处理移除逻辑

    candidate_times = []

    # 如果有偏好时间，优先考虑
    if prefer_time:
        pref_weekday, pref_start, pref_end = prefer_time
        for start_slot in range(pref_start, min(pref_end, 14 - slots_count) + 1):
            if pref_weekday == 4 and start_slot >= 6:  # 周四下午
                continue

            if check_time_available(
                pref_weekday,
                start_slot,
                slots_count,
                classroom_id,
                teacher_ids,
                classes,
                occupied_times,
                weeks=weeks,
            ):
                candidate_times.append((pref_weekday, start_slot, 0))  # 优先级0最高

    # 搜索其他可用时间
    for weekday in range(1, 6):  # 周一到周五
        for start_slot in range(1, 14 - slots_count + 1):
            if weekday == 4 and start_slot >= 6:  # 周四下午
                continue

            # 跳过避免时间
            if avoid_time:
                avoid_weekday, avoid_start, avoid_end = avoid_time
                if weekday == avoid_weekday and not (
                    start_slot + slots_count - 1 < avoid_start or start_slot > avoid_end
                ):
                    continue

            # 如果已经在偏好时间列表中，跳过
            if any(t[0] == weekday and t[1] == start_slot for t in candidate_times):
                continue

            if check_time_available(
                weekday,
                start_slot,
                slots_count,
                classroom_id,
                teacher_ids,
                classes,
                occupied_times,
                weeks=weeks,
            ):
                priority = 1 if (weekday, start_slot) != (old_weekday, old_start) else 2
                candidate_times.append((weekday, start_slot, priority))

    # 注意：不再需要恢复占用，因为我们没有移除

    if candidate_times:
        # 按优先级排序
        candidate_times.sort(key=lambda x: x[2])
        return (candidate_times[0][0], candidate_times[0][1])

    return None


def check_time_available(
    weekday,
    start_slot,
    slots_count,
    classroom_id,
    teacher_ids,
    class_ids,
    occupied_times,
    weeks=None,
):
    """检查指定时间段是否可用（无硬约束冲突）- 考虑周次重叠"""
    for slot in range(start_slot, start_slot + slots_count):
        time_key = (weekday, slot)

        # 检查教室
        if time_key in occupied_times["classroom"][classroom_id]:
            if weeks:
                # 检查周次是否重叠
                for occupied_weeks in occupied_times["classroom"][classroom_id][
                    time_key
                ]:
                    if len(weeks & occupied_weeks) > 0:
                        return False
            else:
                return False  # 无周次信息，保守返回冲突

        # 检查教师
        for tid in teacher_ids:
            if time_key in occupied_times["teacher"][tid]:
                if weeks:
                    for occupied_weeks in occupied_times["teacher"][tid][time_key]:
                        if len(weeks & occupied_weeks) > 0:
                            return False
                else:
                    return False

        # 检查班级
        for cid in class_ids:
            if time_key in occupied_times["class"][cid]:
                if weeks:
                    for occupied_weeks in occupied_times["class"][cid][time_key]:
                        if len(weeks & occupied_weeks) > 0:
                            return False
                else:
                    return False

    return True


def show_remaining_preference_violations(version_id, cursor):
    """显示当前仍未满足的个性化要求详情"""
    from collections import defaultdict

    # 获取当前排课结果（可能已被更新）
    query = """
    SELECT 
        sr.schedule_id,
        sr.task_id,
        sr.classroom_id,
        sr.week_day,
        sr.start_slot,
        tt.slots_count,
        c.course_name,
        GROUP_CONCAT(DISTINCT t.teacher_id SEPARATOR ', ') AS teacher_ids,
        GROUP_CONCAT(DISTINCT t.teacher_name SEPARATOR ', ') AS teacher_names
    FROM schedules sr
    JOIN teaching_tasks tt ON sr.task_id = tt.task_id
    JOIN course_offerings co ON tt.offering_id = co.offering_id
    JOIN courses c ON co.course_id = c.course_id
    LEFT JOIN offering_teachers ot ON co.offering_id = ot.offering_id
    LEFT JOIN teachers t ON ot.teacher_id = t.teacher_id
    WHERE sr.version_id = %s
    GROUP BY sr.schedule_id, sr.task_id, sr.classroom_id, sr.week_day, 
             sr.start_slot, tt.slots_count, c.course_name
    """

    cursor.execute(query, (version_id,))
    current_results = cursor.fetchall()

    # 构建教师实际排课时间
    teacher_schedules = defaultdict(list)
    for result in current_results:
        teacher_ids_str = result.get("teacher_ids", "")
        if not teacher_ids_str:
            continue

        teacher_ids = teacher_ids_str.split(", ")
        start_slot = result["start_slot"]
        end_slot = start_slot + result["slots_count"] - 1
        weekday = result["week_day"]

        for teacher_id in teacher_ids:
            for slot in range(start_slot, end_slot + 1):
                teacher_schedules[teacher_id].append(
                    {
                        "weekday": weekday,
                        "slot": slot,
                        "course_name": result["course_name"],
                    }
                )

    # 获取教师偏好设置
    pref_query = """
    SELECT 
        tp.preference_id,
        tp.offering_id,
        tp.teacher_id,
        tp.preference_type,
        tp.weekday,
        tp.start_slot,
        tp.end_slot,
        tp.penalty_score,
        t.teacher_name
    FROM teacher_preferences tp
    JOIN teachers t ON tp.teacher_id = t.teacher_id
    WHERE tp.offering_id IN (
        SELECT DISTINCT tt.offering_id 
        FROM teaching_tasks tt
        JOIN schedules s ON tt.task_id = s.task_id
        WHERE s.version_id = %s
    )
    ORDER BY tp.penalty_score DESC, tp.preference_type
    """

    cursor.execute(pref_query, (version_id,))
    preferences = cursor.fetchall()

    if not preferences:
        print("\n✓ 没有设置教师个性化要求")
        return

    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    # 分析未满足的个性化要求
    violated_preferred = []  # 未满足的偏好时间
    violated_avoided = []  # 违反的避免时间

    for pref in preferences:
        teacher_id = pref["teacher_id"]
        teacher_name = pref["teacher_name"]
        pref_type = pref["preference_type"]
        weekday = pref["weekday"]
        start_slot = pref["start_slot"]
        end_slot = pref["end_slot"]

        if not weekday or not start_slot or not end_slot:
            continue

        # 检查该教师在该时间段是否有课
        schedule = teacher_schedules.get(teacher_id, [])
        courses_in_range = []

        for sch in schedule:
            if sch["weekday"] == weekday and start_slot <= sch["slot"] <= end_slot:
                if sch["course_name"] not in [c["name"] for c in courses_in_range]:
                    courses_in_range.append(
                        {"name": sch["course_name"], "slot": sch["slot"]}
                    )

        has_class = len(courses_in_range) > 0

        if pref_type == "PREFERRED" and not has_class:
            violated_preferred.append(
                {
                    "teacher_name": teacher_name,
                    "weekday": weekday,
                    "start_slot": start_slot,
                    "end_slot": end_slot,
                    "penalty_score": pref["penalty_score"],
                }
            )
        elif pref_type == "AVOIDED" and has_class:
            violated_avoided.append(
                {
                    "teacher_name": teacher_name,
                    "weekday": weekday,
                    "start_slot": start_slot,
                    "end_slot": end_slot,
                    "penalty_score": pref["penalty_score"],
                    "courses": courses_in_range,
                }
            )

    # 统计
    total_preferred = sum(1 for p in preferences if p["preference_type"] == "PREFERRED")
    total_avoided = sum(1 for p in preferences if p["preference_type"] == "AVOIDED")

    print(f"\n个性化要求满足情况汇总：")
    print(f"  偏好时间（PREFERRED）：")
    print(f"    总计: {total_preferred} 条")
    print(f"    已满足: {total_preferred - len(violated_preferred)} 条")
    print(f"    未满足: {len(violated_preferred)} 条")
    if total_preferred > 0:
        satisfaction_rate = (
            (total_preferred - len(violated_preferred)) / total_preferred * 100
        )
        print(f"    满足率: {satisfaction_rate:.1f}%")

    print(f"\n  避免时间（AVOIDED）：")
    print(f"    总计: {total_avoided} 条")
    print(f"    已遵守: {total_avoided - len(violated_avoided)} 条")
    print(f"    被违反: {len(violated_avoided)} 条")
    if total_avoided > 0:
        compliance_rate = (total_avoided - len(violated_avoided)) / total_avoided * 100
        print(f"    遵守率: {compliance_rate:.1f}%")

    # 显示详细的未满足情况
    if violated_preferred:
        print(f"\n仍未满足的偏好时间详情（共 {len(violated_preferred)} 条）：")
        # 按惩罚分数排序
        violated_preferred.sort(key=lambda x: x["penalty_score"], reverse=True)
        for i, item in enumerate(violated_preferred[:15], 1):  # 显示前15条
            print(f"  [{i}] 教师: {item['teacher_name']}")
            print(
                f"      偏好时间: {day_names[item['weekday']]} 第{item['start_slot']}-{item['end_slot']}节"
            )
            print(f"      惩罚分数: {item['penalty_score']}")
        if len(violated_preferred) > 15:
            print(f"  ... 还有 {len(violated_preferred) - 15} 条未显示")
    else:
        print("\n✓ 所有偏好时间都已满足")

    if violated_avoided:
        print(f"\n仍被违反的避免时间详情（共 {len(violated_avoided)} 条）：")
        # 按惩罚分数排序
        violated_avoided.sort(key=lambda x: x["penalty_score"], reverse=True)
        for i, item in enumerate(violated_avoided[:15], 1):  # 显示前15条
            print(f"  [{i}] 教师: {item['teacher_name']}")
            print(
                f"      避免时间: {day_names[item['weekday']]} 第{item['start_slot']}-{item['end_slot']}节"
            )
            print(f"      惩罚分数: {item['penalty_score']}")
            print(
                f"      冲突课程: {', '.join(set(c['name'] for c in item['courses']))}"
            )
        if len(violated_avoided) > 15:
            print(f"  ... 还有 {len(violated_avoided) - 15} 条未显示")
    else:
        print("\n✓ 所有避免时间都已遵守")

    # 总体评估
    total_violations = len(violated_preferred) + len(violated_avoided)
    total_prefs = total_preferred + total_avoided

    if total_violations == 0:
        print("\n" + "=" * 80)
        print("🎉 所有个性化要求都已完美满足！")
        print("=" * 80)
    else:
        overall_satisfaction = (
            (total_prefs - total_violations) / total_prefs * 100
            if total_prefs > 0
            else 100
        )
        print(
            f"\n总体个性化要求满足率: {overall_satisfaction:.1f}% ({total_prefs - total_violations}/{total_prefs})"
        )

        if total_violations > 0:
            print("\n建议：")
            print("  • 如果硬约束冲突已全部解决，但个性化要求仍未满足，")
            print("    可能是资源限制导致，建议：")
            print("    1. 放宽部分个性化要求的时间范围")
            print("    2. 降低部分个性化要求的优先级（penalty_score）")
            print("    3. 增加可用教室或调整课程时间安排")


def regenerate_conflict_report_after_optimization(version_id, cursor):
    """优化后重新生成冲突报告和导出Excel"""
    from collections import defaultdict

    # 重新获取优化后的排课结果
    query = """
    SELECT 
        sr.schedule_id,
        sr.task_id,
        sr.classroom_id,
        sr.week_day,
        sr.start_slot,
        tt.slots_count,
        tt.offering_id,
        c.course_name,
        cr.classroom_name,
        co.start_week,
        co.end_week,
        co.week_pattern,
        GROUP_CONCAT(DISTINCT t.teacher_name SEPARATOR ', ') AS teacher_name,
        GROUP_CONCAT(DISTINCT t.teacher_id SEPARATOR ', ') AS teacher_ids
    FROM schedules sr
    JOIN teaching_tasks tt ON sr.task_id = tt.task_id
    JOIN course_offerings co ON tt.offering_id = co.offering_id
    JOIN courses c ON co.course_id = c.course_id
    JOIN classrooms cr ON sr.classroom_id = cr.classroom_id
    LEFT JOIN offering_teachers ot ON co.offering_id = ot.offering_id
    LEFT JOIN teachers t ON ot.teacher_id = t.teacher_id
    WHERE sr.version_id = %s
    GROUP BY sr.schedule_id, sr.task_id, sr.classroom_id, sr.week_day, sr.start_slot,
             tt.slots_count, tt.offering_id, c.course_name, cr.classroom_name,
             co.start_week, co.end_week, co.week_pattern
    ORDER BY sr.week_day, sr.start_slot
    """
    cursor.execute(query, (version_id,))
    results = cursor.fetchall()

    # 重新分析冲突（使用简化版本，只分析不执行优化）
    print(f"\n优化后的排课统计：")
    print(f"  - 总课程数: {len(results)}")

    # 分析班级、教师、教室冲突
    class_conflicts_data = []
    teacher_conflicts_data = []
    classroom_conflicts_data = []
    capacity_violations = []

    # 加载offering_weeks数据
    cursor.execute("SELECT offering_id, week_number FROM offering_weeks")
    offering_weeks = {}
    for row in cursor.fetchall():
        offering_id = row["offering_id"]
        week_number = row["week_number"]
        if offering_id not in offering_weeks:
            offering_weeks[offering_id] = set()
        offering_weeks[offering_id].add(week_number)

    def get_weeks(result):
        """根据offering_id获取周次集合"""
        offering_id = result["offering_id"]
        if offering_id in offering_weeks:
            return offering_weeks[offering_id]

        start_week = result.get("start_week") or 1
        end_week = result.get("end_week") or 18
        week_pattern = result.get("week_pattern") or "CONTINUOUS"

        if week_pattern == "CONTINUOUS":
            return set(range(start_week, end_week + 1))
        elif week_pattern == "SINGLE":
            return set(w for w in range(start_week, end_week + 1) if w % 2 == 1)
        elif week_pattern == "DOUBLE":
            return set(w for w in range(start_week, end_week + 1) if w % 2 == 0)
        else:
            return set(range(start_week, end_week + 1))

    # 获取任务-班级关系
    task_classes = defaultdict(list)
    for result in results:
        task_id = result["task_id"]
        class_query = """
        SELECT cl.class_id, cl.class_name
        FROM offering_classes oc
        JOIN classes cl ON oc.class_id = cl.class_id
        JOIN teaching_tasks tt ON oc.offering_id = tt.offering_id
        WHERE tt.task_id = %s
        """
        cursor.execute(class_query, (task_id,))
        classes = cursor.fetchall()
        task_classes[task_id] = classes

    # 检查容量冲突
    for result in results:
        student_query = """
        SELECT SUM(cl.student_count) as total_students
        FROM offering_classes oc
        JOIN classes cl ON oc.class_id = cl.class_id
        JOIN teaching_tasks tt ON oc.offering_id = tt.offering_id
        WHERE tt.task_id = %s
        """
        cursor.execute(student_query, (result["task_id"],))
        student_result = cursor.fetchone()
        student_count = int(student_result["total_students"] or 0)

        classroom_query = "SELECT capacity FROM classrooms WHERE classroom_id = %s"
        cursor.execute(classroom_query, (result["classroom_id"],))
        classroom_result = cursor.fetchone()
        classroom_capacity = (
            int(classroom_result["capacity"]) if classroom_result else 0
        )

        if classroom_capacity < student_count:
            end_slot = result["start_slot"] + result["slots_count"] - 1
            capacity_violations.append(
                {
                    "course": result["course_name"],
                    "classroom": result["classroom_name"],
                    "capacity": classroom_capacity,
                    "students": student_count,
                    "shortage": student_count - classroom_capacity,
                    "weekday": result["week_day"],
                    "start_slot": result["start_slot"],
                    "end_slot": end_slot,
                }
            )

    # 检查时间冲突（简化版）
    class_schedule = defaultdict(lambda: defaultdict(list))
    teacher_schedule = defaultdict(lambda: defaultdict(list))
    classroom_schedule = defaultdict(lambda: defaultdict(list))

    for result in results:
        start_slot = result["start_slot"]
        end_slot = start_slot + result["slots_count"] - 1
        weekday = result["week_day"]
        weeks = get_weeks(result)

        for slot in range(start_slot, end_slot + 1):
            # 班级冲突
            for cls in task_classes[result["task_id"]]:
                class_schedule[cls["class_id"]][(weekday, slot)].append((result, weeks))

            # 教师冲突
            if result["teacher_ids"]:
                for teacher_id in result["teacher_ids"].split(", "):
                    teacher_schedule[teacher_id][(weekday, slot)].append(
                        (result, weeks)
                    )

            # 教室冲突
            classroom_schedule[result["classroom_id"]][(weekday, slot)].append(
                (result, weeks)
            )

    # 统计冲突数量
    def count_conflicts_optimized(schedule_dict):
        """优化后的冲突统计"""
        conflicts = 0
        for entity_id, time_dict in schedule_dict.items():
            for time_key, items in time_dict.items():
                if len(items) > 1:
                    # 检查周次重叠
                    has_overlap = False
                    for i in range(len(items)):
                        for j in range(i + 1, len(items)):
                            weeks1 = items[i][1]
                            weeks2 = items[j][1]
                            if weeks1 & weeks2:  # 有交集
                                has_overlap = True
                                break
                        if has_overlap:
                            break
                    if has_overlap:
                        conflicts += 1
        return conflicts

    class_conflicts = count_conflicts_optimized(class_schedule)
    teacher_conflicts = count_conflicts_optimized(teacher_schedule)
    classroom_conflicts = count_conflicts_optimized(classroom_schedule)

    print(f"  - 班级冲突: {class_conflicts}")
    print(f"  - 教师冲突: {teacher_conflicts}")
    print(f"  - 教室冲突: {classroom_conflicts}")
    print(f"  - 容量不足: {len(capacity_violations)}")

    total_conflicts = (
        class_conflicts
        + teacher_conflicts
        + classroom_conflicts
        + len(capacity_violations)
    )

    if total_conflicts == 0:
        print("\n✅ 优化成功！所有硬冲突已解决")
    else:
        print(f"\n⚠️  仍有 {total_conflicts} 处冲突需要人工处理")

    # 导出优化后的报告（简化版，只包含剩余冲突）
    if capacity_violations:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from datetime import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = "优化后剩余容量冲突"

        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)
        conflict_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )

        day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

        ws.append(
            ["序号", "课程", "星期", "节次", "教室", "教室容量", "学生数", "缺少座位"]
        )
        for col in range(1, 9):
            ws.cell(1, col).fill = header_fill
            ws.cell(1, col).font = header_font
            ws.cell(1, col).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        for idx, item in enumerate(capacity_violations, 1):
            time_range = (
                f"{item['start_slot']}-{item['end_slot']}节"
                if item["start_slot"] != item["end_slot"]
                else f"{item['start_slot']}节"
            )
            ws.append(
                [
                    idx,
                    item["course"],
                    day_names[item["weekday"]],
                    time_range,
                    item["classroom"],
                    item["capacity"],
                    item["students"],
                    item["shortage"],
                ]
            )
            for col in range(1, 9):
                ws.cell(idx + 1, col).fill = conflict_fill

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 12

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"优化后冲突报告_版本{version_id}_{timestamp}.xlsx"
        wb.save(filename)
        print(f"\n✓ 优化后冲突报告已导出到: {filename}")
    else:
        print("\n✓ 无需导出报告（所有容量冲突已解决）")


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("用法: python analyze_conflicts.py <version_id>")
        print("示例: python analyze_conflicts.py 1")
        sys.exit(1)

    version_id = int(sys.argv[1])
    analyze_schedule_conflicts(version_id)
